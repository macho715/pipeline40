# -*- coding: utf-8 -*-
"""
HVDC 입출고 이상치 탐지 — Balanced Boost Edition (v4)
- Rule + Statistical(Per-Location) + ML(IsolationForest/PyOD) 3단 혼합
- **Balanced Boost**: 규칙/통계 근거가 있을 때 ML 위험도를 가중(가산)하여
  허위 양성은 낮추고 진짜 이상은 끌어올림.
- **ECDF(순위) 캘리브레이션 + 베타-스무딩**으로 1.000 포화 방지(0.001~0.999).
- **Per-Location IQR/MAD**: MOSB/DSV… 지점별 정상 체류분포를 따로 학습.
- 헤더 정규화 강화(공백/대소/변형), Excel/JSON 리포트 출력.

설치(선택):
    pip install pyod>=2.0.5 scikit-learn pandas openpyxl xlsxwriter

주의: PyOD 미설치 환경에서도 sklearn IsolationForest로 자동 폴백합니다.
"""
from __future__ import annotations

import json
import logging
import math
from dataclasses import dataclass
from datetime import datetime, timedelta
from enum import Enum
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd

# Optional deps
try:
    from pyod.models.iforest import IForest as PyODIForest  # type: ignore

    PYOD_AVAILABLE = True
except Exception:
    PYOD_AVAILABLE = False

try:
    from sklearn.ensemble import IsolationForest
    from sklearn.preprocessing import StandardScaler

    SKLEARN_AVAILABLE = True
except Exception:
    SKLEARN_AVAILABLE = False

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger("balanced_boost")
logger.setLevel(logging.INFO)
if not logger.handlers:
    logger.addHandler(logging.StreamHandler())


# ----- Domain enums ------------------------------------------------------------
class AnomalyType(Enum):
    DATA_QUALITY = "데이터 품질"
    TIME_REVERSAL = "시간 역전"
    EXCESSIVE_DWELL = "과도 체류"
    ML_OUTLIER = "머신러닝 이상치"


class AnomalySeverity(Enum):
    LOW = "낮음"
    MEDIUM = "보통"
    HIGH = "높음"
    CRITICAL = "치명적"


@dataclass
class AnomalyRecord:
    case_id: str
    anomaly_type: AnomalyType
    severity: AnomalySeverity
    description: str
    detected_value: Optional[float]
    expected_range: Optional[Tuple[float, float]]
    location: Optional[str]
    timestamp: datetime
    risk_score: Optional[float] = None  # [0..1] calibrated

    def to_dict(self) -> Dict:
        return {
            "Case_ID": self.case_id,
            "Anomaly_Type": self.anomaly_type.value,
            "Severity": self.severity.value,
            "Description": self.description,
            "Detected_Value": self.detected_value,
            "Expected_Range": self.expected_range,
            "Location": self.location,
            "Timestamp": self.timestamp.isoformat(),
            "Risk_Score": (
                None if self.risk_score is None else round(float(self.risk_score), 4)
            ),
        }


# ----- Config -----------------------------------------------------------------
@dataclass
class DetectorConfig:
    # 헤더 정규화(동의어 매핑): Master > Slave
    column_map: Dict[str, str] = None
    # 창고/현장 열(정규화된 이름 사용)
    warehouse_columns: List[str] = None
    site_columns: List[str] = None

    # 통계 탐지 파라미터
    iqr_k: float = 1.5
    mad_k: float = 3.5
    min_group_size: int = 10  # 위치별 정상분포 추정 최소 표본

    # ML 탐지 파라미터
    use_pyod_first: bool = True
    contamination: float = 0.02  # 2% 가정(데이터에 따라 조절)
    random_state: int = 42

    # 가중치
    rule_boost: float = 0.25  # 시간역전 발생 시 ML위험도 가산
    stat_boost_high: float = 0.15  # 통계 이상(높음/치명)의 가산치
    stat_boost_med: float = 0.08  # 통계 이상(보통)의 가산치

    # 알림(선택)
    min_risk_to_alert: float = 0.9

    def __post_init__(self):
        if self.column_map is None:
            # Master 헤더 이름으로 정규화
            self.column_map = {
                # 키 필드
                "Case No.": "CASE_NO",
                "CASE NO": "CASE_NO",
                "CASE_NO": "CASE_NO",
                # HVDC CODE
                "HVDC CODE": "HVDC_CODE",
                "HVDC Code": "HVDC_CODE",
                # 창고/현장(표기 변형 흡수)
                "DSV Indoor": "DSV_INDOOR",
                "DSV Al Markaz": "DSV_AL_MARKAZ",
                "AAA Storage": "AAA_STORAGE",
                "AAA  Storage": "AAA_STORAGE",  # double-space 변형
                "DSV Outdoor": "DSV_OUTDOOR",
                "MOSB": "MOSB",
                "Hauler Indoor": "HAULER_INDOOR",
                "DHL Warehouse": "DHL_WAREHOUSE",
                "DSV MZP": "DSV_MZP",
                "AGI": "AGI",
                "DAS": "DAS",
                "MIR": "MIR",
                "SHU": "SHU",
                # 금액/수량 등
                "금액": "AMOUNT",
                "수량": "QTY",
                "Pkg": "PKG",
            }
        if self.warehouse_columns is None:
            self.warehouse_columns = [
                "AAA_STORAGE",
                "DSV_AL_MARKAZ",
                "DSV_INDOOR",
                "DSV_MZP",
                "DSV_OUTDOOR",
                "HAULER_INDOOR",
                "MOSB",
                "DHL_WAREHOUSE",
            ]
        if self.site_columns is None:
            self.site_columns = ["AGI", "DAS", "MIR", "SHU"]


# ----- Utilities ---------------------------------------------------------------
class HeaderNormalizer:
    def __init__(self, column_map: Dict[str, str]):
        self.map = {k.lower(): v for k, v in column_map.items()}

    def normalize(self, df: pd.DataFrame) -> pd.DataFrame:
        new_cols = []
        for c in df.columns:
            key = str(c).strip().lower()
            new_cols.append(self.map.get(key, str(c).strip().upper().replace(" ", "_")))
        df = df.copy()
        df.columns = new_cols
        return df


class DataQualityValidator:
    """간단/빠른 정합성 검증(필요 시 Great Expectations/Pandera로 확장)"""

    HVDC_PATTERN = r"^HVDC-ADOPT-\d{3}-\d{4}$"

    def validate(self, df: pd.DataFrame) -> List[str]:
        issues: List[str] = []
        if "CASE_NO" not in df.columns:
            issues.append("필수 필드 누락: CASE_NO")
        else:
            dup = df["CASE_NO"].astype(str).duplicated().sum()
            if dup:
                issues.append(f"CASE_NO 중복 {dup}건")

        if "HVDC_CODE" in df.columns:
            bad = ~df["HVDC_CODE"].astype(str).str.match(self.HVDC_PATTERN, na=False)
            if bad.any():
                issues.append(f"HVDC_CODE 패턴 불일치 {int(bad.sum())}건")

        # 날짜 변환 가능성(창고/현장)
        warehouse_columns = [
            "AAA_STORAGE",
            "DSV_AL_MARKAZ",
            "DSV_INDOOR",
            "DSV_MZP",
            "DSV_OUTDOOR",
            "HAULER_INDOOR",
            "MOSB",
            "DHL_WAREHOUSE",
        ]
        site_columns = ["AGI", "DAS", "MIR", "SHU"]
        for col in df.columns:
            if col in set(warehouse_columns + site_columns):
                s = pd.to_datetime(df[col], errors="coerce")
                fail_mask = (df[col].notna()) & (s.isna())
                fail = int(fail_mask.sum())
                if fail:
                    issues.append(f"{col}: 날짜 변환 실패 {fail}건")
        return issues


# ----- Feature engineering -----------------------------------------------------
class FeatureBuilder:
    def __init__(self, cfg: DetectorConfig):
        self.cfg = cfg

    def build(
        self, df: pd.DataFrame
    ) -> Tuple[pd.DataFrame, List[Tuple[str, str, int]]]:
        """
        반환:
          - 행 단위 피처(정규화된 CASE_NO index)
          - dwell 목록[(case_id, location, dwell_days)]
        """
        rows = []
        dwell_list: List[Tuple[str, str, int]] = []

        for _, row in df.iterrows():
            case_id = str(row.get("CASE_NO", "NA"))
            points: List[Tuple[str, pd.Timestamp]] = []
            for col in self.cfg.warehouse_columns + self.cfg.site_columns:
                if col in row.index and pd.notna(row[col]):
                    dt = pd.to_datetime(row[col], errors="coerce")
                    if pd.notna(dt):
                        points.append((col, dt))

            points.sort(key=lambda x: x[1])
            if len(points) >= 2:
                # dwell(다음 지점까지 체류일)
                for (loc_a, t_a), (loc_b, t_b) in zip(points[:-1], points[1:]):
                    dd = int((t_b - t_a).days)
                    if dd >= 0:
                        dwell_list.append((case_id, loc_a, dd))

            # 피처(간단형)
            first_ts = points[0][1] if points else pd.NaT
            last_ts = points[-1][1] if points else pd.NaT
            total_days = (
                int((last_ts - first_ts).days)
                if (pd.notna(first_ts) and pd.notna(last_ts))
                else np.nan
            )
            rows.append(
                dict(
                    CASE_NO=case_id,
                    TOUCH_COUNT=len(points),
                    TOTAL_DAYS=total_days,
                    FIRST_TS=first_ts,
                    LAST_TS=last_ts,
                    AMOUNT=pd.to_numeric(row.get("AMOUNT", np.nan), errors="coerce"),
                    QTY=pd.to_numeric(row.get("QTY", np.nan), errors="coerce"),
                    PKG=pd.to_numeric(row.get("PKG", np.nan), errors="coerce"),
                )
            )

        feat = pd.DataFrame(rows).set_index("CASE_NO", drop=True)
        return feat, dwell_list


# ----- Statistical detectors ---------------------------------------------------
class StatDetector:
    def __init__(self, cfg: DetectorConfig):
        self.cfg = cfg

    def per_location_outliers(
        self, dwell_list: List[Tuple[str, str, int]]
    ) -> List[AnomalyRecord]:
        """위치별 IQR(+MAD 보정) 임계로 과도 체류 판정"""
        if not dwell_list:
            return []
        df = pd.DataFrame(dwell_list, columns=["CASE_NO", "LOCATION", "DWELL"])
        out: List[AnomalyRecord] = []

        for loc, g in df.groupby("LOCATION"):
            vals = g["DWELL"].astype(float).values
            if len(vals) < self.cfg.min_group_size:
                continue  # 표본 부족 → 스킵(보수적)

            q1, q3 = np.percentile(vals, 25), np.percentile(vals, 75)
            iqr = max(q3 - q1, 1.0)
            lo_iqr, hi_iqr = q1 - self.cfg.iqr_k * iqr, q3 + self.cfg.iqr_k * iqr

            # MAD 보정(긴 꼬리 방지)
            med = np.median(vals)
            mad = np.median(np.abs(vals - med)) or 1.0
            z = 0.6745 * (vals - med) / mad  # robust z
            # 하한은 공격적이지 않게, 상한만 사용
            hi = max(hi_iqr, med + self.cfg.mad_k * mad)

            for case_id, d in g[["CASE_NO", "DWELL"]].itertuples(index=False):
                if d > hi:
                    # 심각도: hi 초과량 기준
                    ratio = (d - hi) / max(iqr, 1.0)
                    if ratio >= 2.5:
                        sev = AnomalySeverity.CRITICAL
                    elif ratio >= 1.5:
                        sev = AnomalySeverity.HIGH
                    else:
                        sev = AnomalySeverity.MEDIUM

                    out.append(
                        AnomalyRecord(
                            case_id=str(case_id),
                            anomaly_type=AnomalyType.EXCESSIVE_DWELL,
                            severity=sev,
                            description=f"{loc}에서 {int(d)}일 체류 (정상≈{lo_iqr:.1f}~{hi:.1f}일)",
                            detected_value=float(d),
                            expected_range=(float(lo_iqr), float(hi)),
                            location=loc,
                            timestamp=datetime.now(),
                        )
                    )
        return out


# ----- Rule-based detectors ----------------------------------------------------
class RuleDetector:
    def __init__(self, cfg: DetectorConfig):
        self.cfg = cfg

    def time_reversal(self, row: pd.Series) -> Optional[AnomalyRecord]:
        pts: List[Tuple[str, pd.Timestamp]] = []
        for col in self.cfg.warehouse_columns + self.cfg.site_columns:
            if col in row.index and pd.notna(row[col]):
                ts = pd.to_datetime(row[col], errors="coerce")
                if pd.notna(ts):
                    pts.append((col, ts))
        if len(pts) < 2:
            return None

        # 시간 역전이 있는지 확인 (정렬 전후 비교)
        pts_sorted = sorted(pts, key=lambda x: x[1])
        is_same_order = all(
            a[0] == b[0] and a[1] == b[1] for a, b in zip(pts, pts_sorted)
        )
        if not is_same_order:
            # 첫 역전 구간만 설명에 노출
            for (name_a, t_a), (name_b, t_b) in zip(pts[:-1], pts[1:]):
                if t_a > t_b:
                    desc = f"{name_a}({t_a.date()}) → {name_b}({t_b.date()}) 시간 역전"
                    return AnomalyRecord(
                        case_id=str(row.get("CASE_NO", "NA")),
                        anomaly_type=AnomalyType.TIME_REVERSAL,
                        severity=AnomalySeverity.CRITICAL,
                        description=desc,
                        detected_value=None,
                        expected_range=None,
                        location=None,
                        timestamp=datetime.now(),
                        risk_score=0.999,  # balanced 위험도 기준 상한 근사치
                    )
        return None


# ----- Calibration -------------------------------------------------------------
class ECDFCalibrator:
    """
    순위 기반 ECDF를 베타-스무딩으로 0과 1 포화 방지.
    반환값은 (0.001, 0.999) 범위.
    """

    def __init__(self, eps: float = 1e-9):
        self.eps = eps
        self.n: Optional[int] = None
        self.order: Optional[np.ndarray] = None

    def fit(self, raw: np.ndarray) -> "ECDFCalibrator":
        raw = np.asarray(raw, dtype=float)
        self.n = max(len(raw), 1)
        # rank: 1..n (동점 평균)
        from scipy.stats import rankdata  # if unavailable, fallback 아래

        try:
            r = rankdata(raw, method="average")
        except Exception:
            # 간단한 fallback: argsort 기반 평균 순위
            order = np.argsort(raw)
            r = np.empty_like(order, dtype=float)
            r[order] = np.arange(1, len(raw) + 1, dtype=float)
        # 베타-스무딩
        p = (r + 1.0) / (self.n + 2.0)
        self.order = p
        return self

    def transform(self, raw: np.ndarray) -> np.ndarray:
        if self.n is None:
            raise RuntimeError("calibrator is not fit")
        raw = np.asarray(raw, dtype=float)
        if len(raw) == len(self.order):
            p = np.asarray(self.order, dtype=float)
        else:
            # 새 샘플: 분위수 위치를 선형근사
            q = np.argsort(np.argsort(raw)).astype(float) + 1.0
            p = (q + 1.0) / (len(raw) + 2.0)
        # 0,1 포화 방지
        p = np.clip(p, 0.001, 0.999)
        return p


# ----- ML detector -------------------------------------------------------------
class MLDetector:
    def __init__(
        self,
        contamination: float = 0.02,
        random_state: int = 42,
        use_pyod_first: bool = True,
    ):
        self.contamination = contamination
        self.random_state = random_state
        self.use_pyod_first = use_pyod_first and PYOD_AVAILABLE
        self.model = None
        self.scaler = None
        self.calib = ECDFCalibrator()

    def fit_predict(self, X: pd.DataFrame) -> Tuple[np.ndarray, np.ndarray]:
        """return: (y_pred[0/1], risk[0..1])"""
        if X.empty or (not SKLEARN_AVAILABLE and not PYOD_AVAILABLE):
            return np.zeros(len(X), dtype=int), np.zeros(len(X), dtype=float)

        self.scaler = StandardScaler() if SKLEARN_AVAILABLE else None
        Xs = self.scaler.fit_transform(X.values) if self.scaler else X.values

        if self.use_pyod_first:
            # PyOD IForest
            self.model = PyODIForest(
                contamination=self.contamination, random_state=self.random_state
            )
            self.model.fit(Xs)
            # PyOD의 decision_scores_: 값이 클수록 이상치
            raw = np.asarray(self.model.decision_scores_, dtype=float)
            risk = ECDFCalibrator().fit(raw).transform(raw)
            y = (risk >= (1 - self.contamination)).astype(int)
            return y, risk

        # Sklearn IsolationForest (decision_function: 클수록 정상)
        self.model = IsolationForest(
            contamination=self.contamination,
            random_state=self.random_state,
            n_estimators=256,
        )
        self.model.fit(Xs)
        dec = self.model.decision_function(Xs)  # +: 정상, -: 이상
        # 위험도 = 1 - ECDF(dec)
        risk = 1.0 - ECDFCalibrator().fit(dec).transform(dec)
        y = (risk >= (1 - self.contamination)).astype(int)
        return y, risk


# ----- Balanced Boost combiner -------------------------------------------------
class BalancedCombiner:
    """규칙/통계 신호를 ML 위험도에 가산."""

    def __init__(self, cfg: DetectorConfig):
        self.cfg = cfg
        self.rule_cases: set[str] = set()
        self.stat_cases_med: set[str] = set()
        self.stat_cases_high: set[str] = set()

    def ingest_rule(self, recs: Iterable[AnomalyRecord]):
        for r in recs:
            if r.anomaly_type == AnomalyType.TIME_REVERSAL:
                self.rule_cases.add(r.case_id)

    def ingest_stat(self, recs: Iterable[AnomalyRecord]):
        for r in recs:
            if r.severity in (AnomalySeverity.HIGH, AnomalySeverity.CRITICAL):
                self.stat_cases_high.add(r.case_id)
            else:
                self.stat_cases_med.add(r.case_id)

    def fuse(self, case_id: str, ml_risk: float) -> float:
        r = float(ml_risk)
        if case_id in self.rule_cases:
            r += self.cfg.rule_boost
        if case_id in self.stat_cases_high:
            r += self.cfg.stat_boost_high
        elif case_id in self.stat_cases_med:
            r += self.cfg.stat_boost_med
        return float(np.clip(r, 0.001, 0.999))


# ----- Orchestrator ------------------------------------------------------------
class HybridAnomalyDetector:
    def __init__(self, cfg: DetectorConfig):
        self.cfg = cfg
        self.normalizer = HeaderNormalizer(cfg.column_map)
        self.validator = DataQualityValidator()
        self.rule = RuleDetector(cfg)
        self.stat = StatDetector(cfg)
        self.ml = MLDetector(cfg.contamination, cfg.random_state, cfg.use_pyod_first)
        self.comb = BalancedCombiner(cfg)

    def run(
        self,
        df_raw: pd.DataFrame,
        export_excel: Optional[str] = None,
        export_json: Optional[str] = None,
    ) -> Dict:
        df = self.normalizer.normalize(df_raw)
        issues = self.validator.validate(df)
        anomalies: List[AnomalyRecord] = []
        if issues:
            logger.warning(f"데이터 품질 이슈: {issues}")
            anomalies.extend(
                [
                    AnomalyRecord(
                        case_id=(
                            str(df.iloc[0].get("CASE_NO", "NA")) if len(df) else "NA"
                        ),
                        anomaly_type=AnomalyType.DATA_QUALITY,
                        severity=AnomalySeverity.MEDIUM,
                        description="; ".join(issues),
                        detected_value=None,
                        expected_range=None,
                        location=None,
                        timestamp=datetime.now(),
                        risk_score=None,
                    )
                ]
            )

        # Rule — row-wise
        for _, row in df.iterrows():
            ar = self.rule.time_reversal(row)
            if ar:
                anomalies.append(ar)

        # Features & Dwell
        feat, dwell_list = FeatureBuilder(self.cfg).build(df)

        # Statistical — per location
        stat_recs = self.stat.per_location_outliers(dwell_list)
        anomalies.extend(stat_recs)

        # Balanced: feed to combiner
        self.comb.ingest_rule(
            [a for a in anomalies if a.anomaly_type == AnomalyType.TIME_REVERSAL]
        )
        self.comb.ingest_stat(stat_recs)

        # ML
        use_cols = [
            c
            for c in ["TOUCH_COUNT", "TOTAL_DAYS", "AMOUNT", "QTY", "PKG"]
            if c in feat.columns
        ]
        X = feat[use_cols].fillna(0.0)
        y, risk = self.ml.fit_predict(X)

        if len(X):
            for case_id, yi, ri in zip(X.index, y, risk):
                if yi == 1:
                    fused = self.comb.fuse(str(case_id), float(ri))
                    sev = (
                        AnomalySeverity.CRITICAL
                        if fused >= 0.97
                        else (
                            AnomalySeverity.HIGH
                            if fused >= 0.90
                            else AnomalySeverity.MEDIUM
                        )
                    )
                    anomalies.append(
                        AnomalyRecord(
                            case_id=str(case_id),
                            anomaly_type=AnomalyType.ML_OUTLIER,
                            severity=sev,
                            description=f"ML 이상치(위험도 {fused:.3f})",
                            detected_value=float(fused),
                            expected_range=None,
                            location=None,
                            timestamp=datetime.now(),
                            risk_score=float(fused),
                        )
                    )

        # Summary & export
        summary = self._build_summary(anomalies)
        if export_json:
            self._export_json(Path(export_json), anomalies)
        if export_excel:
            # 지표 덤프 포함
            self._export_excel(Path(export_excel), anomalies, feat.reset_index())
        return {"summary": summary, "count": len(anomalies), "anomalies": anomalies}

    # -------- Summary/Exporters --------
    def _build_summary(self, anomalies: List[AnomalyRecord]) -> Dict:
        by_type: Dict[str, int] = {}
        by_sev: Dict[str, int] = {}
        for a in anomalies:
            by_type[a.anomaly_type.value] = by_type.get(a.anomaly_type.value, 0) + 1
            by_sev[a.severity.value] = by_sev.get(a.severity.value, 0) + 1
        return {"total": len(anomalies), "by_type": by_type, "by_severity": by_sev}

    def _export_json(self, path: Path, anomalies: List[AnomalyRecord]) -> None:
        data = [a.to_dict() for a in anomalies]
        path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        logger.info(f"JSON 저장: {path}")

    def _export_excel(
        self, path: Path, anomalies: List[AnomalyRecord], feat: pd.DataFrame
    ) -> None:
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl 미설치로 Excel 생략")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["총계", sum(1 for _ in anomalies)])
        ws.append([])
        from collections import Counter

        ct_type = Counter(a.anomaly_type.value for a in anomalies)
        ct_sev = Counter(a.severity.value for a in anomalies)
        ws.append(["유형", "건수"])
        for k, v in ct_type.items():
            ws.append([k, v])
        ws.append([])
        ws.append(["심각도", "건수"])
        for k, v in ct_sev.items():
            ws.append([k, v])

        # 상세
        ws2 = wb.create_sheet("Anomalies")
        ws2.append(
            [
                "Case_ID",
                "Anomaly_Type",
                "Severity",
                "Description",
                "Detected_Value",
                "Expected_Range",
                "Location",
                "Timestamp",
                "Risk_Score",
            ]
        )
        for a in anomalies:
            d = a.to_dict()
            ws2.append(
                [
                    d["Case_ID"],
                    d["Anomaly_Type"],
                    d["Severity"],
                    d["Description"],
                    d["Detected_Value"],
                    str(d["Expected_Range"]),
                    d["Location"],
                    d["Timestamp"],
                    d["Risk_Score"],
                ]
            )

        # 피처
        ws3 = wb.create_sheet("Features")
        for i, col in enumerate(feat.columns, start=1):
            ws3.cell(row=1, column=i).value = col
        for r, (_, row) in enumerate(feat.iterrows(), start=2):
            for c, col in enumerate(feat.columns, start=1):
                ws3.cell(row=r, column=c).value = row[col]

        wb.save(path)
        logger.info(f"Excel 저장: {path}")


# ----- CLI (optional) ----------------------------------------------------------
def _load_excel(path: str, sheet: Optional[str] = None) -> pd.DataFrame:
    if path.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return pd.read_excel(path, sheet_name=sheet)
    return pd.read_csv(path)


def main():
    import argparse

    p = argparse.ArgumentParser(description="HVDC 이상치 탐지 — Balanced Boost Edition")
    p.add_argument("--input", required=True, help="Stage3 Excel 또는 CSV")
    p.add_argument("--sheet", default=None, help="엑셀 시트명(기본 자동)")
    p.add_argument("--out-json", default="HVDC_anomaly_report_balanced.json")
    p.add_argument("--out-xlsx", default="HVDC_anomaly_report_balanced.xlsx")
    p.add_argument("--contamination", type=float, default=0.02)
    p.add_argument("--use-pyod", action="store_true", help="가능하면 PyOD 사용")
    args = p.parse_args()

    cfg = DetectorConfig(contamination=args.contamination, use_pyod_first=args.use_pyod)
    df = _load_excel(args.input, args.sheet)

    det = HybridAnomalyDetector(cfg)
    res = det.run(df, export_json=args.out_json, export_excel=args.out_xlsx)
    logger.info(res)


if __name__ == "__main__":
    main()
