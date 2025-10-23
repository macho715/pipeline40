#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HVDC 이상치 보고서에 통합_원본데이터_Fixed 시트 추가 및 색상 적용
"""

import json
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from typing import Dict, List


def create_final_colored_report():
    """
    HVDC_anomaly_report.xlsx에 통합_원본데이터_Fixed 시트를 추가하고
    이상치가 발견된 케이스에 색상 마킹을 적용합니다.
    """
    
    print("=" * 80)
    print("HVDC 이상치 보고서 최종 색상 적용")
    print("=" * 80)
    print()
    
    # 1. Stage 3 보고서에서 통합_원본데이터_Fixed 로드
    print("[Step 1] Stage 3 보고서에서 통합_원본데이터_Fixed 시트 로드...")
    stage3_reports = list(Path("data/processed/reports").glob("HVDC_*.xlsx"))
    if not stage3_reports:
        print("ERROR: Stage 3 보고서를 찾을 수 없습니다.")
        return False
    
    latest_report = sorted(stage3_reports, key=lambda p: p.stat().st_mtime)[-1]
    print(f"   최신 보고서: {latest_report.name}")
    
    try:
        df_source = pd.read_excel(latest_report, sheet_name="통합_원본데이터_Fixed")
        print(f"   데이터 로드 완료: {len(df_source)}행 × {len(df_source.columns)}열")
    except Exception as e:
        print(f"ERROR: 시트 로드 실패: {e}")
        return False
    
    # 2. HVDC_anomaly_report.xlsx 로드
    print("\n[Step 2] HVDC_anomaly_report.xlsx 로드...")
    anomaly_report_path = Path("data/anomaly/HVDC_anomaly_report.xlsx")
    if not anomaly_report_path.exists():
        print(f"ERROR: {anomaly_report_path}를 찾을 수 없습니다.")
        return False
    
    wb = openpyxl.load_workbook(anomaly_report_path)
    print(f"   기존 시트: {', '.join(wb.sheetnames)}")
    
    # 3. "통합_원본데이터_Fixed" 시트 추가 (기존 시트가 있으면 삭제)
    print("\n[Step 3] 통합_원본데이터_Fixed 시트 추가...")
    if "통합_원본데이터_Fixed" in wb.sheetnames:
        print("   기존 시트 삭제 중...")
        del wb["통합_원본데이터_Fixed"]
    
    ws = wb.create_sheet("통합_원본데이터_Fixed", 0)  # 첫 번째 위치에 추가
    print("   시트 생성 완료")
    
    # 4. DataFrame을 시트에 쓰기
    print("   데이터 쓰기 중...")
    for r_idx, row in enumerate(dataframe_to_rows(df_source, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    print(f"   데이터 쓰기 완료: {ws.max_row}행")
    
    # 5. 이상치 JSON 로드
    print("\n[Step 4] 이상치 JSON 데이터 로드...")
    json_path = Path("data/anomaly/HVDC_anomaly_report.json")
    if not json_path.exists():
        print(f"ERROR: {json_path}를 찾을 수 없습니다.")
        return False
    
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    # JSON 구조 확인: dict with 'anomalies' key or list
    if isinstance(data, dict) and "anomalies" in data:
        anomalies = data["anomalies"]
    elif isinstance(data, list):
        anomalies = data
    else:
        print(f"ERROR: Unexpected JSON structure: {type(data)}")
        return False
    
    print(f"   이상치 로드 완료: {len(anomalies)}건")
    
    # 6. 색상 정의
    colors = {
        "시간 역전": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),  # 빨강
        "머신러닝 이상치": {
            "치명적": PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),  # 주황
            "높음": PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
            "보통": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),  # 노랑
            "낮음": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
        },
        "데이터 품질": PatternFill(start_color="FFCC99FF", end_color="FFCC99FF", fill_type="solid"),  # 보라
        "과도 체류": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),  # 노랑
        "최종위치 불일치": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),  # 노랑
    }
    
    # 7. Case NO 컬럼 찾기
    print("\n[Step 5] Case NO 컬럼 찾기...")
    case_col = None
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value and (str(header_value) == "Case No." or "CASE_NO" in str(header_value).upper()):
            case_col = col
            print(f"   Case NO 컬럼 발견: {col}번째 컬럼 ({header_value})")
            break
    
    if not case_col:
        print("ERROR: Case NO 컬럼을 찾을 수 없습니다.")
        return False
    
    # 8. 색상 적용
    print("\n[Step 6] 이상치 색상 적용 중...")
    applied_count = 0
    anomaly_counts = {
        "시간 역전": 0,
        "ml_high": 0,
        "ml_medium": 0,
        "데이터 품질": 0,
        "기타": 0
    }
    
    # Case ID → Row 매핑 생성 (성능 향상)
    case_id_map = {}
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=case_col).value
        if cell_value:
            case_id_map[str(cell_value).strip()] = row
    
    print(f"   Case ID 매핑 완료: {len(case_id_map)}개")
    
    for i, anomaly in enumerate(anomalies, 1):
        if i % 500 == 0:
            print(f"   처리 중: {i}/{len(anomalies)}...")
        
        # 키 이름 정규화 (Case_ID, case_id 모두 지원)
        case_id = str(anomaly.get("case_id") or anomaly.get("Case_ID") or "").strip()
        anom_type = anomaly.get("anomaly_type") or anomaly.get("Anomaly_Type") or ""
        severity = anomaly.get("severity") or anomaly.get("Severity") or ""
        
        # Case ID 매칭
        if case_id not in case_id_map:
            continue
        
        row_idx = case_id_map[case_id]
        
        # 색상 선택
        fill = None
        if "시간 역전" in anom_type or "TIME_REVERSAL" in anom_type:
            fill = colors["시간 역전"]
            anomaly_counts["시간 역전"] += 1
            # 날짜 컬럼만 색칠
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header and any(kw in str(header).lower() for kw in ["date", "날짜", "time", "시간", "warehouse", "site", "dhl", "dsv", "agi", "das", "mir", "shu"]):
                    ws.cell(row=row_idx, column=col).fill = fill
            applied_count += 1
            continue
        
        elif "머신러닝" in anom_type or "ML_OUTLIER" in anom_type:
            if severity in ["치명적", "높음", "CRITICAL", "HIGH"]:
                fill = colors["머신러닝 이상치"]["높음"]
                anomaly_counts["ml_high"] += 1
            else:
                fill = colors["머신러닝 이상치"]["보통"]
                anomaly_counts["ml_medium"] += 1
        
        elif "데이터 품질" in anom_type or "DATA_QUALITY" in anom_type:
            fill = colors["데이터 품질"]
            anomaly_counts["데이터 품질"] += 1
        
        elif "과도 체류" in anom_type or "EXCESSIVE_DWELL" in anom_type:
            fill = colors["과도 체류"]
            anomaly_counts["ml_medium"] += 1
        
        elif "최종위치 불일치" in anom_type or "STATUS_MISMATCH" in anom_type:
            fill = colors["최종위치 불일치"]
            anomaly_counts["ml_medium"] += 1
        
        else:
            fill = colors["머신러닝 이상치"]["보통"]
            anomaly_counts["기타"] += 1
        
        # 전체 행 색칠
        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = fill
            applied_count += 1
    
    print(f"   색상 적용 완료: {applied_count}건")
    print(f"   - 시간 역전: {anomaly_counts['시간 역전']}건 (빨강)")
    print(f"   - ML 이상치 (높음): {anomaly_counts['ml_high']}건 (주황)")
    print(f"   - ML 이상치 (보통): {anomaly_counts['ml_medium']}건 (노랑)")
    print(f"   - 데이터 품질: {anomaly_counts['데이터 품질']}건 (보라)")
    print(f"   - 기타: {anomaly_counts['기타']}건")
    
    # 9. 색상 범례 시트 추가
    print("\n[Step 7] 색상 범례 시트 추가...")
    if "색상_범례" in wb.sheetnames:
        del wb["색상_범례"]
    
    legend_sheet = wb.create_sheet("색상_범례")
    
    legend_data = [
        ["색상", "의미", "적용 범위", "개수"],
        ["🔴 빨간색", "시간 역전 이상치", "날짜 컬럼만", str(anomaly_counts["시간 역전"])],
        ["🟠 주황색", "ML 이상치 (높음/치명적)", "전체 행", str(anomaly_counts["ml_high"])],
        ["🟡 노란색", "ML 이상치 (보통/낮음)", "전체 행", str(anomaly_counts["ml_medium"])],
        ["🟣 보라색", "데이터 품질 이상", "전체 행", str(anomaly_counts["데이터 품질"])],
        ["", "", "총 적용", str(applied_count)],
    ]
    
    for row_data in legend_data:
        legend_sheet.append(row_data)
    
    # 범례 시트 서식
    for row in range(1, len(legend_data) + 1):
        for col in range(1, len(legend_data[0]) + 1):
            cell = legend_sheet.cell(row=row, column=col)
            if row == 1:
                cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
    
    # 10. 저장
    print("\n[Step 8] 파일 저장 중...")
    wb.save(anomaly_report_path)
    print(f"   저장 완료: {anomaly_report_path}")
    
    print("\n" + "=" * 80)
    print("✅ 이상치 보고서 최종 색상 적용 완료!")
    print("=" * 80)
    print(f"\n최종 파일: {anomaly_report_path}")
    print(f"시트 목록: {', '.join(wb.sheetnames)}")
    print(f"색상 적용: {applied_count}/{len(anomalies)}건")
    
    return True


if __name__ == "__main__":
    try:
        success = create_final_colored_report()
        exit(0 if success else 1)
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        exit(1)

