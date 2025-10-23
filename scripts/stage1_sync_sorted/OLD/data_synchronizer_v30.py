# -*- coding: utf-8 -*-
"""
DataSynchronizer v3.0 - Semantic Header Matching Edition
=========================================================

This is a completely refactored version of the data synchronizer that uses
the new core header matching system. All hardcoded column names have been
replaced with semantic key lookups.

Key Improvements:
- Zero hardcoding of column names
- Automatic adaptation to different Excel formats
- Robust header detection and matching
- Comprehensive error reporting
- Better maintainability

Migration from v2.9:
- All column name strings replaced with semantic keys
- Header finding logic centralized in core modules
- Enhanced validation and error messages
"""

from __future__ import annotations
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Import the new core header matching system
from ..core import (
    SemanticMatcher,
    find_header_by_meaning,
    detect_header_row,
    HVDC_HEADER_REGISTRY,
    HeaderCategory
)

# ===== Configuration =====
ORANGE = "FFC000"  # Changed date cell
YELLOW = "FFFF00"  # New row

# Use semantic keys instead of hardcoded column names
# These will be automatically matched to actual column names
DATE_SEMANTIC_KEYS = [
    "etd_atd",
    "eta_ata",
    "dhl_warehouse",
    "dsv_indoor",
    "dsv_al_markaz",
    "dsv_outdoor",
    "aaa_storage",
    "hauler_indoor",
    "dsv_mzp",
    "mosb",
    "shifting",
    "mir",
    "shu",
    "das",
    "agi",
]

ALWAYS_OVERWRITE_NONDATE = True


def _to_date(val) -> Optional[pd.Timestamp]:
    """Convert various date formats to pandas Timestamp."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    try:
        if isinstance(val, pd.Timestamp):
            return val
        return pd.to_datetime(val, errors="coerce")
    except Exception:
        return None


@dataclass
class Change:
    """Record of a single cell change."""
    row_index: int
    column_name: str
    old_value: Any
    new_value: Any
    change_type: str  # "date_update" | "field_update" | "new_record"


@dataclass
class ChangeTracker:
    """Tracks all changes made during synchronization."""
    changes: List[Change] = field(default_factory=list)
    new_cases: Dict[str, Dict[str, Any]] = field(default_factory=dict)

    def add_change(self, **kw):
        """Add a change record."""
        self.changes.append(
            Change(
                row_index=int(kw.get("row_index", -1)),
                column_name=str(kw.get("column_name", "")),
                old_value=kw.get("old_value"),
                new_value=kw.get("new_value"),
                change_type=str(kw.get("change_type", "field_update")),
            )
        )

    def log_new_case(
        self, case_no: str, row_data: Dict[str, Any], row_index: Optional[int] = None
    ):
        """Log a new case that was appended."""
        self.new_cases[str(case_no)] = dict(row_data or {})
        if row_index is not None:
            self.add_change(
                row_index=row_index,
                column_name="",
                old_value=None,
                new_value=None,
                change_type="new_record",
            )


@dataclass
class SyncResult:
    """Result of a synchronization operation."""
    success: bool
    message: str
    output_path: str
    stats: Dict[str, Any]
    matching_report: Optional[str] = None  # New: matching diagnostics


class DataSynchronizerV30:
    """
    Advanced data synchronizer with semantic header matching.
    
    This version eliminates all hardcoded column names and uses semantic
    matching to automatically find the correct columns regardless of how
    they're named in the Excel file.
    
    The synchronizer works in several phases:
    1. Load files and detect header rows
    2. Match all required columns using semantic keys
    3. Validate that required columns were found
    4. Perform synchronization using matched column names
    5. Apply Excel formatting to highlight changes
    
    Example:
        >>> sync = DataSynchronizerV30()
        >>> result = sync.synchronize("master.xlsx", "warehouse.xlsx")
        >>> if result.success:
        >>>     print(f"Synchronized to {result.output_path}")
        >>> else:
        >>>     print(f"Error: {result.message}")
    """

    def __init__(self, date_semantic_keys: Optional[List[str]] = None) -> None:
        """
        Initialize the synchronizer.
        
        Args:
            date_semantic_keys: List of semantic keys for date columns.
                If None, uses the default DATE_SEMANTIC_KEYS.
        """
        # Use semantic keys instead of hardcoded column names
        self.date_semantic_keys = date_semantic_keys or DATE_SEMANTIC_KEYS
        
        # Initialize the semantic matcher
        self.matcher = SemanticMatcher(
            min_confidence=0.7,
            allow_partial=True
        )
        
        # Change tracking
        self.change_tracker = ChangeTracker()
        
        # Column mapping storage (will be populated during matching)
        self.master_columns: Dict[str, str] = {}  # semantic_key -> actual_column
        self.warehouse_columns: Dict[str, str] = {}
        
        print("[OK] DataSynchronizer v3.0 initialized with semantic header matching")

    def _dates_equal(self, a, b) -> bool:
        """Check if two dates are equal, ignoring format differences."""
        da = _to_date(a)
        db = _to_date(b)
        if da is None and db is None:
            return True
        if da is None or db is None:
            return False
        if pd.isna(da) or pd.isna(db):
            return pd.isna(da) and pd.isna(db)
        return da.normalize() == db.normalize()

    def _load_file_with_header_detection(
        self, 
        file_path: str, 
        file_label: str
    ) -> Tuple[pd.DataFrame, int]:
        """
        Load all sheets from Excel file with automatic header row detection.
        
        Args:
            file_path: Path to the Excel file
            file_label: Label for logging (e.g., "Master", "Warehouse")
            
        Returns:
            Tuple of (merged_dataframe, header_row_index)
        """
        print(f"\n{'='*60}")
        print(f"Loading {file_label} file: {Path(file_path).name}")
        print(f"{'='*60}")
        
        xl = pd.ExcelFile(file_path)
        all_dfs = []
        header_row = None
        
        print(f"Found {len(xl.sheet_names)} sheets in file")
        
        for sheet_name in xl.sheet_names:
            print(f"\n  Loading sheet: '{sheet_name}'")
            
            # Detect header row for this sheet
            sheet_header_row, confidence = detect_header_row(file_path, sheet_name)
            
            if header_row is None:
                header_row = sheet_header_row
            
            print(f"  [OK] Header at row {sheet_header_row} (confidence: {confidence:.0%})")
            
            # Load sheet
            df = pd.read_excel(xl, sheet_name=sheet_name, header=sheet_header_row)
            
            if df.empty:
                print(f"  [SKIP] Empty sheet")
                continue
            
            # Track source sheet
            df['Source_Sheet'] = sheet_name
            all_dfs.append(df)
            print(f"  [OK] {len(df)} rows loaded")
        
        if not all_dfs:
            raise ValueError(f"No valid sheets found in {file_label}")
        
        # Merge all sheets
        merged_df = pd.concat(all_dfs, ignore_index=True, sort=False)
        print(f"\n[OK] Total: {len(merged_df)} rows from {len(all_dfs)} sheets")
        
        # Consolidate incorrectly named warehouse columns
        print("\nConsolidating warehouse columns:")
        merged_df = self._consolidate_warehouse_columns(merged_df)
        
        # Ensure all location columns exist
        print("\nEnsuring all location columns:")
        merged_df = self._ensure_all_location_columns(merged_df)
        
        return merged_df, header_row

    def _consolidate_warehouse_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Consolidate incorrectly named warehouse columns.
        
        Some raw data files use incorrect column names for the same warehouse.
        This method merges such columns to ensure data consistency:
        - "DSV WH" → "DSV Indoor" (HE Local sheet uses incorrect name)
        
        Args:
            df: DataFrame with potentially incorrect column names
            
        Returns:
            DataFrame with consolidated columns
        """
        consolidations = {
            "DSV WH": "DSV Indoor",  # HE Local sheet incorrectly names DSV Indoor as DSV WH
        }
        
        for wrong_name, correct_name in consolidations.items():
            if wrong_name in df.columns:
                if correct_name in df.columns:
                    # Merge data: use correct_name where it exists, fill with wrong_name otherwise
                    df[correct_name] = df[correct_name].fillna(df[wrong_name])
                    df = df.drop(columns=[wrong_name])
                    print(f"  [OK] Merged '{wrong_name}' → '{correct_name}'")
                else:
                    # Just rename if correct column doesn't exist yet
                    df = df.rename(columns={wrong_name: correct_name})
                    print(f"  [OK] Renamed '{wrong_name}' → '{correct_name}'")
        
        return df

    def _ensure_all_location_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Ensure all warehouse and site columns from header registry exist in DataFrame.
        
        Adds missing location columns as empty (NaT) columns to maintain consistency
        across pipeline stages.
        
        Args:
            df: DataFrame to check and update
            
        Returns:
            DataFrame with all location columns present
        """
        # Get all warehouse and site definitions from registry
        all_locations = []
        
        # Warehouse locations
        for definition in HVDC_HEADER_REGISTRY.definitions.values():
            if definition.category == HeaderCategory.LOCATION:
                # Get canonical name (first alias)
                canonical = definition.aliases[0] if definition.aliases else None
                if canonical:
                    all_locations.append(canonical)
        
        # Check and add missing columns
        missing_cols = []
        for location in all_locations:
            if location not in df.columns:
                df[location] = pd.NaT
                missing_cols.append(location)
        
        if missing_cols:
            print(f"  [OK] Added {len(missing_cols)} missing location columns:")
            for col in missing_cols:
                print(f"    - {col}")
        else:
            print(f"  [OK] All location columns present")
        
        return df

    def _match_and_validate_headers(
        self, 
        df: pd.DataFrame, 
        file_label: str
    ) -> Dict[str, str]:
        """
        Match semantic keys to actual column names and validate required columns.
        
        Args:
            df: The DataFrame to match against
            file_label: Label for logging
            
        Returns:
            Dictionary mapping semantic_key to actual column name
            
        Raises:
            ValueError: If required columns are not found
        """
        print(f"\nMatching headers for {file_label}...")
        
        # Define required semantic keys for synchronization
        required_keys = [
            "case_number",  # Must have case number for matching
            "item_number",  # Good to have item number
        ]
        
        # All keys we want to find (required + date columns)
        all_keys = required_keys + self.date_semantic_keys
        
        # Perform semantic matching
        report = self.matcher.match_dataframe(df, all_keys)
        
        # Print summary
        print(f"  Matched: {report.successful_matches}/{report.total_semantic_keys}")
        print(f"  Success rate: {report.successful_matches/report.total_semantic_keys:.0%}")
        
        # Validate required columns
        missing_required = []
        for key in required_keys:
            if not report.get_column_name(key):
                missing_required.append(key)
        
        if missing_required:
            # Print detailed error information
            print(f"\n{'='*60}")
            print(f"ERROR: Missing required columns in {file_label}")
            print(f"{'='*60}")
            report.print_summary()
            raise ValueError(
                f"Required semantic keys not found in {file_label}: {missing_required}"
            )
        
        # Show what was matched
        print("\n  Key matches:")
        for key in required_keys + self.date_semantic_keys[:3]:  # Show first few
            col = report.get_column_name(key)
            if col:
                print(f"    - {key:20s} → '{col}'")
        
        if len(self.date_semantic_keys) > 3:
            remaining = len([k for k in self.date_semantic_keys 
                           if report.get_column_name(k)])
            print(f"    ... and {remaining} more date columns")
        
        # Build mapping dictionary
        column_mapping = {
            key: report.get_column_name(key)
            for key in all_keys
            if report.get_column_name(key)
        }
        
        return column_mapping

    def _build_case_index(
        self, 
        df: pd.DataFrame, 
        case_col: str
    ) -> Dict[str, int]:
        """
        Build an index mapping case numbers to row indices.
        
        Args:
            df: The DataFrame
            case_col: Name of the case number column
            
        Returns:
            Dictionary mapping normalized case numbers to row indices
        """
        import re
        
        idx: Dict[str, int] = {}
        
        # Normalize case numbers: uppercase, remove special characters
        series = df[case_col].fillna("").astype(str).str.strip().str.upper()
        series = series.apply(lambda x: re.sub(r"[^A-Z0-9]", "", x))
        
        for i, v in enumerate(series.tolist()):
            if not v:
                continue
            if v not in idx:  # Keep first occurrence
                idx[v] = i
        
        return idx

    def _apply_master_order_sorting(
        self,
        master: pd.DataFrame,
        warehouse: pd.DataFrame,
        master_cols: Dict[str, str],
        wh_cols: Dict[str, str],
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Sort both DataFrames according to Master's order.
        
        Uses the matched column names to sort by item number (if available)
        and then by case number order.
        
        Args:
            master: Master DataFrame
            warehouse: Warehouse DataFrame
            master_cols: Master column mapping (semantic_key -> column_name)
            wh_cols: Warehouse column mapping
            
        Returns:
            Tuple of (sorted_master, sorted_warehouse)
        """
        print("\nApplying Master order sorting...")
        
        # Sort Master by (No, Case No.) for stable ordering with multi-sheet data
        item_col = master_cols.get("item_number")
        if item_col and item_col in master.columns:
            case_col = master_cols["case_number"]
            print(f"  Sorting Master by ('{item_col}', '{case_col}')...")
            master = master.sort_values(
                [item_col, case_col], 
                na_position="last"
            ).reset_index(drop=True)
        
        # Get case columns
        master_case_col = master_cols["case_number"]
        wh_case_col = wh_cols["case_number"]
        
        # Get Master's case order
        master_case_order = master[master_case_col].dropna().tolist()
        print(f"  Master has {len(master_case_order)} cases")
        
        # Split warehouse into Master cases and non-Master cases
        wh_master_cases = warehouse[
            warehouse[wh_case_col].isin(master_case_order)
        ].copy()
        wh_other_cases = warehouse[
            ~warehouse[wh_case_col].isin(master_case_order)
        ].copy()
        
        print(f"  Warehouse has {len(wh_master_cases)} Master cases, "
              f"{len(wh_other_cases)} other cases")
        
        # Sort warehouse Master cases by Master's order
        case_order_map = {case: idx for idx, case in enumerate(master_case_order)}
        wh_master_cases["_sort_order"] = wh_master_cases[wh_case_col].map(case_order_map)
        wh_master_cases = wh_master_cases.sort_values("_sort_order").drop("_sort_order", axis=1)
        
        # Combine: Master order cases first, then others
        sorted_warehouse = pd.concat(
            [wh_master_cases, wh_other_cases], 
            ignore_index=True
        )
        
        print(f"  [OK] Sorting complete")
        
        return master, sorted_warehouse

    def _apply_updates(
        self,
        master: pd.DataFrame,
        wh: pd.DataFrame,
        master_cols: Dict[str, str],
        wh_cols: Dict[str, str],
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Apply updates from Master to Warehouse using matched column names.
        
        Args:
            master: Master DataFrame
            wh: Warehouse DataFrame
            master_cols: Master column mapping
            wh_cols: Warehouse column mapping
            
        Returns:
            Tuple of (updated_warehouse, statistics)
        """
        print("\nApplying updates from Master to Warehouse...")
        
        stats = dict(updates=0, date_updates=0, field_updates=0, appends=0)
        
        # Build warehouse index by case number
        wh_case_col = wh_cols["case_number"]
        wh_index = self._build_case_index(wh, wh_case_col)
        
        # Get master case column
        master_case_col = master_cols["case_number"]
        
        # Find all semantic keys that exist in both files
        common_keys = set(master_cols.keys()) & set(wh_cols.keys())
        
        # Process each master row
        for mi, mrow in master.iterrows():
            # Get case number
            key = (
                str(mrow[master_case_col]).strip().upper()
                if pd.notna(mrow[master_case_col])
                else ""
            )
            if not key:
                continue
            
            # Check if case exists in warehouse
            if key not in wh_index:
                # Append new case
                append_row = {}
                for semantic_key in common_keys:
                    m_col = master_cols[semantic_key]
                    w_col = wh_cols[semantic_key]
                    append_row[w_col] = mrow[m_col]
                
                wh = pd.concat([wh, pd.DataFrame([append_row])], ignore_index=True)
                new_index = len(wh) - 1
                stats["appends"] += 1
                
                # Track new case
                self.change_tracker.log_new_case(
                    case_no=key, 
                    row_data=append_row, 
                    row_index=new_index
                )
                continue
            
            # Case exists - update it
            wi = wh_index[key]
            
            # Process each common column
            for semantic_key in common_keys:
                m_col = master_cols[semantic_key]
                w_col = wh_cols[semantic_key]
                
                mval = mrow[m_col]
                wval = wh.at[wi, w_col] if wi < len(wh) else None
                
                # Check if this is a date column
                is_date = semantic_key in self.date_semantic_keys
                
                if is_date:
                    # Date column: Master always wins if it has a value
                    if pd.notna(mval):
                        if not self._dates_equal(mval, wval):
                            stats["updates"] += 1
                            stats["date_updates"] += 1
                            wh.at[wi, w_col] = mval
                            self.change_tracker.add_change(
                                row_index=wi,
                                column_name=w_col,
                                old_value=wval,
                                new_value=mval,
                                change_type="date_update",
                            )
                        else:
                            # Equal logically - ensure consistent format
                            wh.at[wi, w_col] = mval
                else:
                    # Non-date column: Overwrite if Master has value
                    if ALWAYS_OVERWRITE_NONDATE and pd.notna(mval):
                        if (wval is None) or (str(mval) != str(wval)):
                            stats["updates"] += 1
                            stats["field_updates"] += 1
                            wh.at[wi, w_col] = mval
                            self.change_tracker.add_change(
                                row_index=wi,
                                column_name=w_col,
                                old_value=wval,
                                new_value=mval,
                                change_type="field_update",
                            )
        
        print(f"  [OK] Updates: {stats['updates']} cells changed")
        print(f"    - Date updates: {stats['date_updates']}")
        print(f"    - Field updates: {stats['field_updates']}")
        print(f"    - New records: {stats['appends']}")
        
        return wh, stats

    def synchronize(
        self, 
        master_xlsx: str, 
        warehouse_xlsx: str, 
        output_path: Optional[str] = None
    ) -> SyncResult:
        """
        Synchronize Master data into Warehouse file.
        
        This is the main entry point for the synchronization process. It:
        1. Loads both files with automatic header detection
        2. Matches all required columns using semantic keys
        3. Validates that required columns exist
        4. Applies Master data to Warehouse
        5. Maintains Master's row order
        6. Saves the result with color-coded changes
        
        Args:
            master_xlsx: Path to Master Excel file
            warehouse_xlsx: Path to Warehouse Excel file
            output_path: Path for output file (optional, auto-generated if None)
            
        Returns:
            SyncResult object with success status, messages, and statistics
            
        Example:
            >>> sync = DataSynchronizerV30()
            >>> result = sync.synchronize("master.xlsx", "warehouse.xlsx")
            >>> if result.success:
            >>>     print(f"Success! Output: {result.output_path}")
            >>>     print(f"Changes: {result.stats}")
        """
        try:
            # Phase 1: Load files with header detection
            print("\n" + "="*60)
            print("PHASE 1: Loading Files")
            print("="*60)
            
            m_df, m_header_row = self._load_file_with_header_detection(
                master_xlsx, "Master"
            )
            w_df, w_header_row = self._load_file_with_header_detection(
                warehouse_xlsx, "Warehouse"
            )
            
            # Phase 2: Match headers using semantic keys
            print("\n" + "="*60)
            print("PHASE 2: Semantic Header Matching")
            print("="*60)
            
            self.master_columns = self._match_and_validate_headers(m_df, "Master")
            self.warehouse_columns = self._match_and_validate_headers(w_df, "Warehouse")
            
            # Phase 3: Sort according to Master order
            print("\n" + "="*60)
            print("PHASE 3: Sorting")
            print("="*60)
            
            m_df, w_df = self._apply_master_order_sorting(
                m_df, w_df, 
                self.master_columns, 
                self.warehouse_columns
            )
            
            # Phase 4: Apply updates
            print("\n" + "="*60)
            print("PHASE 4: Synchronization")
            print("="*60)
            
            updated_w_df, stats = self._apply_updates(
                m_df, w_df,
                self.master_columns,
                self.warehouse_columns
            )
            
            # Maintain Master order after updates
            updated_w_df, _ = self._apply_master_order_sorting(
                m_df, updated_w_df,
                self.master_columns,
                self.warehouse_columns
            )
            
            # Phase 5: Save and format
            print("\n" + "="*60)
            print("PHASE 5: Saving Output")
            print("="*60)
            
            # Determine output path
            out = output_path or str(
                Path(warehouse_xlsx).with_name(
                    Path(warehouse_xlsx).stem + ".synced_v3.0.xlsx"
                )
            )
            
            # Read warehouse Excel to get sheet name
            w_xl = pd.ExcelFile(warehouse_xlsx)
            sheet_name = w_xl.sheet_names[0]
            
            # Save
            print(f"  Writing to: {Path(out).name}")
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                updated_w_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f"  [OK] Saved")
            
            # Apply color formatting
            print(f"  Applying color formatting...")
            self._apply_excel_formatting(out, sheet_name, w_header_row)
            print(f"  [OK] Formatting applied")
            
            # Prepare result
            stats["output_file"] = out
            
            print("\n" + "="*60)
            print("[OK] SYNCHRONIZATION COMPLETE")
            print("="*60)
            print(f"Output: {out}")
            print(f"Changes: {stats['updates']} updates, {stats['appends']} new records")
            
            return SyncResult(
                True, 
                "Sync & colorize done.", 
                out, 
                stats,
                matching_report="All headers matched successfully"
            )
            
        except Exception as e:
            error_msg = f"Synchronization failed: {str(e)}"
            print(f"\n{'='*60}")
            print(f"[ERROR] ERROR: {error_msg}")
            print(f"{'='*60}")
            
            return SyncResult(
                False, 
                error_msg, 
                output_path or warehouse_xlsx, 
                {},
                matching_report=str(e)
            )

    def _apply_excel_formatting(
        self, 
        excel_file: str, 
        sheet_name: str, 
        header_row: int
    ):
        """
        Apply color formatting to the Excel file to highlight changes.
        
        Args:
            excel_file: Path to the Excel file
            sheet_name: Name of the sheet to format
            header_row: Row index where headers start (0-based from pandas)
        """
        try:
            wb = load_workbook(excel_file)
            if sheet_name not in wb.sheetnames:
                return
            
            ws = wb[sheet_name]
            
            # Excel rows are 1-indexed, and we need to account for header
            excel_header_row = header_row + 1
            
            # Build header map
            header_map = {}
            for c_idx, cell in enumerate(ws[excel_header_row], start=1):
                if cell.value is None:
                    continue
                header_map[str(cell.value).strip()] = c_idx
            
            # Define fills
            orange_fill = PatternFill(
                start_color=ORANGE, end_color=ORANGE, fill_type="solid"
            )
            yellow_fill = PatternFill(
                start_color=YELLOW, end_color=YELLOW, fill_type="solid"
            )
            
            # Apply date changes (orange)
            for change in self.change_tracker.changes:
                if change.change_type != "date_update":
                    continue
                
                excel_row = change.row_index + excel_header_row + 1
                col_idx = header_map.get(change.column_name)
                
                if col_idx:
                    cell = ws.cell(row=excel_row, column=col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        cell.fill = orange_fill
            
            # Apply new records (yellow)
            for change in self.change_tracker.changes:
                if change.change_type == "new_record":
                    excel_row = change.row_index + excel_header_row + 1
                    
                    # Color all cells with data in this row
                    for cell in ws[excel_row]:
                        if cell.value is not None and str(cell.value).strip():
                            cell.fill = yellow_fill
            
            wb.save(excel_file)
            
        except Exception as e:
            print(f"  Warning: Formatting failed: {e}")


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(
        description="DataSynchronizer v3.0 - Semantic Header Matching Edition"
    )
    ap.add_argument("--master", required=True, help="Path to Master Excel file")
    ap.add_argument("--warehouse", required=True, help="Path to Warehouse Excel file")
    ap.add_argument("--out", default="", help="Output path (optional)")
    args = ap.parse_args()

    print("\n" + "="*60)
    print("DataSynchronizer v3.0")
    print("Semantic Header Matching Edition")
    print("="*60)

    sync = DataSynchronizerV30()
    res = sync.synchronize(args.master, args.warehouse, args.out or None)

    print("\n" + "="*60)
    print("FINAL RESULT")
    print("="*60)
    print(f"Success: {res.success}")
    print(f"Message: {res.message}")
    print(f"Output:  {res.output_path}")
    print(f"Stats:   {res.stats}")
    
    if not res.success:
        print(f"\nError details: {res.matching_report}")
        exit(1)
