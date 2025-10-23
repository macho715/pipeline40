
# -*- coding: utf-8 -*-
"""
Minimal AnomalyVisualizer
- run_pipeline.py 가 기대하는 인터페이스와 동일
- 지정한 시트에서 case_col 기준으로 이상치 행을 색상표시(노랑/주황/빨강)
"""
from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from shutil import copy2

class AnomalyVisualizer:
    def __init__(self, anomalies: List[Dict[str, Any]]):
        self.anomalies = anomalies or []

    def _color_for(self, severity: str) -> str:
        s = str(severity).strip()
        if "치명" in s or s.lower().startswith("critical"):
            return "FFC7CE"  # red-ish
        if "높음" in s or s.lower().startswith("high"):
            return "FFEB9C"  # yellow
        return "C6E0B4"  # green-ish for medium/low

    def apply_anomaly_colors(
        self,
        *,
        excel_file: str,
        sheet_name: str,
        case_col: str = "Case No.",
        create_backup: bool = True,
    ) -> Dict[str, Any]:
        p = Path(excel_file)
        if not p.exists():
            return {"success": False, "message": f"Excel 파일이 존재하지 않습니다: {p}"}

        if p.suffix.lower() not in {".xlsx",".xlsm"}:
            return {"success": False, "message": "Excel(xlsx/xlsm) 파일만 지원됩니다."}

        try:
            xls = pd.ExcelFile(p)
            if sheet_name not in xls.sheet_names:
                return {"success": False, "message": f"시트를 찾을 수 없습니다: {sheet_name}. 사용가능: {xls.sheet_names}"}
        except Exception as e:
            return {"success": False, "message": f"시트 정보 조회 실패: {e}"}

        target = p
        backup_path = None
        if create_backup:
            backup_path = str(p.with_suffix(".backup.xlsx"))
            copy2(str(p), backup_path)

        wb = load_workbook(str(target))
        ws = wb[sheet_name]

        # 헤더 행 탐색(1~5행에서 case_col과 같은 셀을 찾음)
        header_row_idx = None
        for r in range(1, min(6, ws.max_row+1)):
            values = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
            if case_col in values:
                header_row_idx = r
                break
        if header_row_idx is None:
            return {"success": False, "message": f"헤더 행에서 '{case_col}' 를 찾지 못했습니다."}

        # 케이스 인덱스 매핑
        col_index = [str(c.value).strip() if c.value is not None else "" for c in ws[header_row_idx]].index(case_col) + 1
        case_to_rows: Dict[str, List[int]] = {}
        for r in range(header_row_idx+1, ws.max_row+1):
            v = ws.cell(row=r, column=col_index).value
            if v is None:
                continue
            case_to_rows.setdefault(str(v), []).append(r)

        # 색상 적용
        colored = 0
        for a in self.anomalies:
            case_id = str(a.get("case_id",""))
            sev = a.get("severity","")
            fill = PatternFill("solid", fgColor=self._color_for(str(sev)))
            for r in case_to_rows.get(case_id, []):
                for c in ws[r]:
                    c.fill = fill
                colored += 1

        wb.save(str(target))
        return {"success": True, "message": f"{colored}개 행에 색상 적용", "backup_path": backup_path}
