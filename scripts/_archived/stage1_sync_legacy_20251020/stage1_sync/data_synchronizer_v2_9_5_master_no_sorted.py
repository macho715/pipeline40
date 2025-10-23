#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Data Synchronizer — FULL Patched (v2.9.5, Master-NO ordered)
- Master always takes precedence
- _norm_key (Master/Warehouse unified)
- Multi-update for duplicate Case No. (update all rows)
- Union-update (auto-create missing Master columns in Warehouse)
- Robust date normalization to YYYY-MM-DD (AM/PM, HHMMSS, Excel serial)
- Safe output path (no /data permission issue; fallback to warehouse dir)
- NEW: Output sorted by MASTER 'NO' (fallback to MASTER row order if NO missing)
- Index remap after sorting so highlights stay correct
- Excel save with fixed date format + built-in minimal colorizer
"""

from __future__ import annotations

import argparse
import math
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from datetime import datetime

# -------------------------------
# Config & Constants
# -------------------------------

ORANGE = "FFC000"  # date update
YELLOW = "FFFF00"  # new row

DATE_KEYS = [
    "ETD/ATD",
    "ETA/ATA",
    "DHL Warehouse",
    "DSV Indoor",
    "DSV Al Markaz",
    "DSV Outdoor",
    "AAA  Storage",
    "Hauler Indoor",
    "DSV MZP",
    "MOSB",
    "Shifting",
    "MIR",
    "SHU",
    "DAS",
    "AGI",
]

ALWAYS_OVERWRITE_NONDATE = True
CLEAR_IF_MASTER_EMPTY_DATES = True

try:
    PROJECT_ROOT = Path(__file__).resolve().parent
except Exception:
    PROJECT_ROOT = Path.cwd()


# -------------------------------
# Utilities
# -------------------------------

def _norm_header(s: str) -> str:
    """Normalize header for matching: strip, lower, collapse spaces/underscores."""
    if s is None:
        return ""
    s2 = str(s).strip().lower()
    s2 = re.sub(r"[\s_]+", " ", s2)
    return s2


def _is_date_col(col: str) -> bool:
    return _norm_header(col) in {_norm_header(k) for k in DATE_KEYS}


def _norm_key(val: Any) -> str:
    """CASE 키 정규화: NFKC→대문자→비영숫자 제거→정수화 처리 (e.g., 208221.0 → 208221)."""
    if val is None:
        return ""
    try:
        if isinstance(val, float) and math.isnan(val):
            return ""
    except Exception:
        pass
    s = unicodedata.normalize("NFKC", str(val)).strip().upper()
    # float-like "208221.0" → "208221" (fraction must be zero)
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s


_AMPM = re.compile(r"^\s*\d{4}-\d{2}-\d{2}\s+\d{1,2}:\d{2}:\d{2}\s*(am|pm)\s*$", re.I)
_AMPM2 = re.compile(r"^\s*\d{4}-\d{2}-\d{2}\s+\d{1,2}:\d{2}\s*(am|pm)\s*$", re.I)
_TIME6 = re.compile(r"^\s*\d{4}-\d{2}-\d{2}\s+\d{6}\s*$")  # 2024-10-03 190000
_TIME4 = re.compile(r"^\s*\d{4}-\d{2}-\d{2}\s+\d{4}\s*$")  # 2024-10-03 1900


def _normalize_date(val) -> pd.Timestamp | pd.NaT:
    """Normalize to date only (YYYY-MM-DD). Accepts AM/PM, HHMMSS, Excel serial, common strings."""
    if val is None:
        return pd.NaT
    # pandas NA
    if isinstance(val, float) and np.isnan(val):
        return pd.NaT

    # Excel serial (int/float days since 1899-12-30)
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            base = pd.Timestamp("1899-12-30")
            ts = base + pd.to_timedelta(float(val), unit="D")
            return pd.Timestamp(ts.date())
        except Exception:
            pass

    s = str(val).strip()
    if not s:
        return pd.NaT
    s = re.sub(r"\s+", " ", s)

    # 2025-05-13 12:00:00 AM
    if _AMPM.match(s):
        try:
            ts = datetime.strptime(s, "%Y-%m-%d %I:%M:%S %p")
            return pd.Timestamp(ts.date())
        except Exception:
            pass
    # 2025-05-13 12:00 AM
    if _AMPM2.match(s):
        try:
            ts = datetime.strptime(s, "%Y-%m-%d %I:%M %p")
            return pd.Timestamp(ts.date())
        except Exception:
            pass
    # 2024-10-03 190000
    if _TIME6.match(s):
        try:
            ts = datetime.strptime(s, "%Y-%m-%d %H%M%S")
            return pd.Timestamp(ts.date())
        except Exception:
            pass
    # 2024-10-03 1900
    if _TIME4.match(s):
        try:
            ts = datetime.strptime(s, "%Y-%m-%d %H%M")
            return pd.Timestamp(ts.date())
        except Exception:
            pass

    # Fallback: pandas parse
    try:
        ts = pd.to_datetime(s, errors="raise", dayfirst=False, utc=False)
        return pd.Timestamp(ts.date())
    except Exception:
        return pd.NaT


def _normalize_dataframe_dates(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize all DATE_KEYS columns to date-only."""
    if df is None or df.empty:
        return df
    date_norms = {_norm_header(k) for k in DATE_KEYS}
    for col in df.columns:
        if _norm_header(col) in date_norms:
            df[col] = df[col].map(_normalize_date)
    return df


def _dates_equal(a, b) -> bool:
    da, db = _normalize_date(a), _normalize_date(b)
    if pd.isna(da) and pd.isna(db):
        return True
    if pd.isna(da) or pd.isna(db):
        return False
    return da == db


def resolve_synced_output_path(
    warehouse_xlsx: str | Path,
    *, project_root: Optional[Path] = None,
) -> Path:
    """
    Resolve synchronized output path (safe; avoid /data). Fallback to warehouse folder if permission denied.
    """
    warehouse_xlsx = Path(warehouse_xlsx).resolve()
    default_root = warehouse_xlsx.parent
    root = Path(project_root) if project_root else (default_root or PROJECT_ROOT)

    synced_dir = root / "data/processed/synced"
    if not synced_dir.exists():
        try:
            synced_dir.mkdir(parents=True, exist_ok=True)
        except PermissionError:
            synced_dir = default_root
            synced_dir.mkdir(parents=True, exist_ok=True)

    base = warehouse_xlsx.stem
    output_name = f"{base}.synced.xlsx"
    return synced_dir / output_name


# -------------------------------
# Change tracker (minimal)
# -------------------------------

class ChangeTracker:
    def __init__(self) -> None:
        self.changes: List[Dict[str, Any]] = []
        self.new_cases: List[Dict[str, Any]] = []

    def add_change(self, *, row_index: int, column_name: str, old_value: Any, new_value: Any, change_type: str) -> None:
        self.changes.append(dict(row_index=row_index, column=column_name, old=old_value, new=new_value, type=change_type))

    def log_new_case(self, *, case_no: str, row_data: Dict[str, Any], row_index: int) -> None:
        self.new_cases.append(dict(case_no=case_no, row=row_data, row_index=row_index))


# -------------------------------
# Result dataclass
# -------------------------------

@dataclass
class SyncResult:
    ok: bool
    message: str
    output_path: Path
    stats: Dict[str, Any]


# -------------------------------
# Core Synchronizer
# -------------------------------

class DataSynchronizerV29:
    def __init__(self, date_keys: Optional[List[str]] = None) -> None:
        self.date_keys = date_keys or DATE_KEYS
        self.change_tracker = ChangeTracker()

    # ---- column detection
    def _case_col(self, df: pd.DataFrame) -> Optional[str]:
        cands = ["Case No.", "CASE NO", "case_no", "CASE NO.", "Case No", "CASE"]
        norm_set = {_norm_header(c) for c in df.columns}
        for c in cands:
            if _norm_header(c) in norm_set:
                for col in df.columns:
                    if _norm_header(col) == _norm_header(c):
                        return col
        # fallback to first column that contains 'case'
        for col in df.columns:
            if "case" in str(col).lower():
                return col
        return None

    def _no_col(self, df: pd.DataFrame) -> Optional[str]:
        cands = ["NO", "No", "no", "NO.", "번호"]
        norm_set = {_norm_header(c) for c in df.columns}
        for c in cands:
            if _norm_header(c) in norm_set:
                for col in df.columns:
                    if _norm_header(col) == _norm_header(c):
                        return col
        return None

    # ---- index: multi-map for duplicate Case No., normalized with _norm_key
    def _build_index(self, df: pd.DataFrame, case_col: str) -> Dict[str, List[int]]:
        idx: Dict[str, List[int]] = {}
        series = df[case_col].map(_norm_key)
        for i, v in enumerate(series.tolist()):
            if not v:
                continue
            idx.setdefault(v, []).append(i)
        return idx

    # ---- union-update support: ensure missing master columns exist in warehouse
    def _ensure_wh_has_master_columns(self, master: pd.DataFrame, wh: pd.DataFrame) -> pd.DataFrame:
        m_norm_map = {_norm_header(c): c for c in master.columns}
        w_norm_set = {_norm_header(c) for c in wh.columns}
        for k, src in m_norm_map.items():
            if k not in w_norm_set:
                wh[src] = pd.NaT if _is_date_col(src) else pd.NA
        return wh

    # ---- sorting by MASTER NO (fallback to MASTER row order)
    def _sort_wh_by_master_no(self, master: pd.DataFrame, wh: pd.DataFrame, case_col_m: str, case_col_w: str) -> tuple[pd.DataFrame, Dict[int, int], Dict[str, Any]]:
        no_col = self._no_col(master)

        temp = master.copy()
        temp["_key"] = temp[case_col_m].map(_norm_key)
        temp["_row_rank"] = np.arange(len(temp))
        temp["_no"] = pd.to_numeric(temp[no_col], errors="coerce") if no_col else np.nan

        order_map: Dict[str, Tuple[float, int]] = {}
        for _, r in temp[["_key", "_no", "_row_rank"]].iterrows():
            k = r["_key"]
            if not k:
                continue
            no_key = float(r["_no"]) if pd.notna(r["_no"]) else float("inf")
            if k not in order_map:
                order_map[k] = (no_key, int(r["_row_rank"]))

        # Build sort keys for WH rows
        keys = wh[case_col_w].map(_norm_key).tolist()
        tuple_keys: List[Tuple[float, int, int, int]] = []
        dup_seq: Dict[str, int] = {}
        for idx, k in enumerate(keys):
            base = order_map.get(k, (float("inf"), 10_000_000 + idx))
            dup_seq[k] = dup_seq.get(k, 0) + 1
            tuple_keys.append((*base, dup_seq[k], idx))

        order = np.argsort(np.array(tuple_keys, dtype=object), kind="mergesort")
        wh_sorted = wh.iloc[order].reset_index(drop=True)

        # map old_index -> new_index
        old_to_new: Dict[int, int] = {int(old): int(new) for new, old in enumerate(order.tolist())}
        meta = dict(used_no=bool(no_col), no_column=no_col or "", rows_reordered=int(np.sum(order != np.arange(len(order)))))
        return wh_sorted, old_to_new, meta

    def _remap_change_indices(self, old_to_new: Dict[int, int]) -> None:
        for ch in self.change_tracker.changes:
            oi = ch.get("row_index", None)
            if isinstance(oi, int) and oi in old_to_new:
                ch["row_index"] = old_to_new[oi]
        for it in self.change_tracker.new_cases:
            oi = it.get("row_index", None)
            if isinstance(oi, int) and oi in old_to_new:
                it["row_index"] = old_to_new[oi]

    # ---- updates
    def _apply_updates(self, master: pd.DataFrame, wh: pd.DataFrame, case_col_m: str, case_col_w: str) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        stats: Dict[str, Any] = dict(updates=0, date_updates=0, field_updates=0, appends=0, clears=0)

        # normalize dates first for stability
        master = _normalize_dataframe_dates(master.copy())
        wh = _normalize_dataframe_dates(wh.copy())

        wh = self._ensure_wh_has_master_columns(master, wh)

        wh_index = self._build_index(wh, case_col_w)
        dup_cases = {k: v for k, v in wh_index.items() if len(v) > 1}
        stats["warehouse_duplicate_cases"] = len(dup_cases)

        m_norm = {_norm_header(c): c for c in master.columns}
        w_norm = {_norm_header(c): c for c in wh.columns}
        aligned = [(m_norm[k], w_norm[k]) for k in sorted(set(m_norm) & set(w_norm))]

        for mi, mrow in master.iterrows():
            key = _norm_key(mrow.get(case_col_m, ""))
            if not key:
                continue

            if key not in wh_index:
                # append new row with aligned columns
                append_row = {wcol: mrow[mcol] for (mcol, wcol) in aligned}
                wh = pd.concat([wh, pd.DataFrame([append_row])], ignore_index=True)
                new_index = len(wh) - 1
                wh_index[key] = wh_index.get(key, []) + [new_index]
                stats["appends"] += 1
                self.change_tracker.log_new_case(case_no=key, row_data=append_row, row_index=new_index)
                continue

            target_rows = wh_index[key]
            for wi in target_rows:
                for mcol, wcol in aligned:
                    mval = mrow[mcol]
                    wval = wh.at[wi, wcol] if (wi < len(wh) and wcol in wh.columns) else None
                    is_date = _is_date_col(wcol)

                    if is_date:
                        m_norm_val = _normalize_date(mval)
                        w_norm_val = _normalize_date(wval)

                        if pd.notna(m_norm_val):
                            if not _dates_equal(m_norm_val, w_norm_val):
                                stats["updates"] += 1; stats["date_updates"] += 1
                                wh.at[wi, wcol] = m_norm_val
                                self.change_tracker.add_change(row_index=wi, column_name=wcol, old_value=wval, new_value=str(pd.Timestamp(m_norm_val).date()), change_type="date_update")
                            else:
                                # re-write normalized value for format stability
                                wh.at[wi, wcol] = m_norm_val
                        else:
                            if CLEAR_IF_MASTER_EMPTY_DATES and pd.notna(w_norm_val):
                                wh.at[wi, wcol] = pd.NaT
                                stats["updates"] += 1; stats["date_updates"] += 1; stats["clears"] += 1
                                self.change_tracker.add_change(row_index=wi, column_name=wcol, old_value=wval, new_value=None, change_type="date_update")
                    else:
                        if ALWAYS_OVERWRITE_NONDATE and pd.notna(mval):
                            if (wval is None) or (str(mval) != str(wval)):
                                stats["updates"] += 1; stats["field_updates"] += 1
                                wh.at[wi, wcol] = mval
                                self.change_tracker.add_change(row_index=wi, column_name=wcol, old_value=wval, new_value=mval, change_type="field_update")

        return wh, stats

    # ---- main
    def synchronize(self, master_xlsx: str | Path, warehouse_xlsx: str | Path, output_path: Optional[str | Path] = None) -> SyncResult:
        try:
            master_xlsx = Path(master_xlsx).resolve()
            warehouse_xlsx = Path(warehouse_xlsx).resolve()

            m_xl = pd.ExcelFile(master_xlsx)
            w_xl = pd.ExcelFile(warehouse_xlsx)

            m_df = pd.read_excel(m_xl, sheet_name=m_xl.sheet_names[0])
            w_df = pd.read_excel(w_xl, sheet_name=w_xl.sheet_names[0])

            m_case = self._case_col(m_df)
            w_case = self._case_col(w_df)
            if not (m_case and w_case):
                return SyncResult(False, "CASE NO column not found.", Path(output_path) if output_path else warehouse_xlsx, {})

            # union columns + updates (includes date normalization)
            updated_w_df, stats = self._apply_updates(m_df, w_df, m_case, w_case)

            # NEW: sort by MASTER NO (or MASTER row order)
            sorted_w_df, idx_map, sort_meta = self._sort_wh_by_master_no(m_df, updated_w_df, m_case, w_case)
            self._remap_change_indices(idx_map)
            stats.update(dict(sorted_by_master_no=sort_meta["used_no"], no_column=sort_meta["no_column"], rows_reordered=sort_meta["rows_reordered"]))

            # resolve output path
            out_path = Path(output_path) if output_path else resolve_synced_output_path(warehouse_xlsx)

            # save with fixed date format + in-place coloring
            sheet_name = w_xl.sheet_names[0]
            with pd.ExcelWriter(out_path, engine="openpyxl", datetime_format="YYYY-MM-DD", date_format="YYYY-MM-DD") as xw:
                sorted_w_df.to_excel(xw, index=False, sheet_name=sheet_name)

                # minimal colorizer (no external import dependency)
                try:
                    from openpyxl.styles import PatternFill
                    wb = xw.book
                    ws = wb[sheet_name]

                    header_map = {str(c.value).strip(): idx+1 for idx, c in enumerate(ws[1]) if c.value is not None}
                    orange = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
                    yellow = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

                    # new rows
                    for it in self.change_tracker.new_cases:
                        r = it.get("row_index", None)
                        if isinstance(r, int):
                            excel_row = r + 2
                            for cell in ws[excel_row]:
                                cell.fill = yellow

                    # updated cells
                    for ch in self.change_tracker.changes:
                        r = ch.get("row_index", None)
                        col = ch.get("column", ch.get("column_name"))
                        typ = ch.get("type", ch.get("change_type"))
                        if not isinstance(r, int) or not col:
                            continue
                        excel_row = r + 2
                        col_idx = header_map.get(col)
                        if col_idx is None:
                            key = str(col).strip().lower().replace(" ", "_")
                            for k, v in header_map.items():
                                if key == str(k).strip().lower().replace(" ", "_"):
                                    col_idx = v
                                    break
                        if col_idx is None:
                            continue
                        if typ == "new_record":
                            for cell in ws[excel_row]:
                                cell.fill = yellow
                        else:
                            ws.cell(row=excel_row, column=col_idx).fill = orange
                except Exception:
                    pass

            msg = "Sync & save done."
            return SyncResult(True, msg, out_path, stats)

        except Exception as e:
            return SyncResult(False, f"Error: {e}", Path(output_path) if output_path else Path("."), {})


# -------------------------------
# CLI
# -------------------------------

def main():
    p = argparse.ArgumentParser(description="CASE MASTER → WAREHOUSE Synchronizer (v2.9.5, Master-NO ordered)")
    p.add_argument("--master", required=True, help="Path to CASE MASTER Excel (Case List.xlsx)")
    p.add_argument("--warehouse", required=True, help="Path to WAREHOUSE Excel (HVDC WAREHOUSE_*.xlsx)")
    p.add_argument("--out", default=None, help="Output Excel path (optional; if omitted, auto path is used)")
    args = p.parse_args()

    res = DataSynchronizerV29().synchronize(args.master, args.warehouse, args.out)
    print("ok:", res.ok)
    print("message:", res.message)
    print("output:", res.output_path)
    print("stats:", res.stats)


if __name__ == "__main__":
    main()
