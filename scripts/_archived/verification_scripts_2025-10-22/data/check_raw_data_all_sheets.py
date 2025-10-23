"""
RAW DATA 파일의 모든 시트와 컬럼을 상세히 확인하는 스크립트
"""

import pandas as pd
import openpyxl


def check_file(file_name, file_path):
    print(f"\n{'='*80}")
    print(f"파일: {file_name}")
    print(f"경로: {file_path}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(file_path)
    print(f"총 {len(wb.sheetnames)}개 시트 발견\n")

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"\n[시트 {sheet_idx}/{len(wb.sheetnames)}: {sheet_name}]")
        print("-" * 80)

        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            print(f"총 행 수: {len(df):,}행")
            print(f"총 컬럼 수: {len(df.columns)}개\n")

            print("컬럼 목록:")
            for i, col in enumerate(df.columns, 1):
                data_count = df[col].notna().sum()
                percentage = (data_count / len(df) * 100) if len(df) > 0 else 0

                # DHL 관련 컬럼 하이라이트
                highlight = " ★ DHL 관련" if "DHL" in str(col).upper() else ""

                print(f"  {i:3d}. [{col}]")
                print(f"       데이터: {data_count:,}건 ({percentage:.1f}%){highlight}")

                # DHL 관련 컬럼이면 샘플 데이터 출력
                if "DHL" in str(col).upper() and data_count > 0:
                    sample = df[df[col].notna()][col].head(3).tolist()
                    print(f"       샘플: {sample}")

            print()

        except Exception as e:
            print(f"  오류: {e}\n")


# 메인 실행
print("\n" + "=" * 80)
print("RAW DATA 전체 시트 상세 분석")
print("=" * 80)

files = [
    ("CASE LIST.xlsx", "data/raw/CASE LIST.xlsx"),
    ("HVDC WAREHOUSE_HITACHI(HE).xlsx", "data/raw/HVDC WAREHOUSE_HITACHI(HE).xlsx"),
]

for file_name, file_path in files:
    check_file(file_name, file_path)

print("\n" + "=" * 80)
print("분석 완료")
print("=" * 80)
