#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Comprehensive color verification for all stages"""

import sys
import io
import openpyxl
from pathlib import Path
from collections import Counter

# UTF-8 encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def check_stage1():
    """Check Stage 1 synced file colors"""
    print("="*80)
    print("Stage 1 색상 검증")
    print("="*80)
    
    file_path = Path("data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v2.9.4.xlsx")
    if not file_path.exists():
        print(f"[ERROR] 파일 없음: {file_path}")
        return False
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print(f"파일: {file_path.name}")
    print(f"시트: {ws.title}, {ws.max_row}행 x {ws.max_column}열")
    
    colors = Counter()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                c = str(cell.fill.fgColor.rgb)
                if c not in ['00000000', 'None']:
                    colors[c] += 1
    
    print(f"\n색상 결과:")
    if colors:
        for color, count in colors.most_common():
            name = "알 수 없음"
            if color in ["FFC000", "FFFFC000", "00FFC000"]:
                name = "[주황] 날짜 변경"
            elif color in ["FFFF00", "FFFFFF00", "00FFFF00"]:
                name = "[노랑] 신규 행"
            print(f"  {name} ({color}): {count}개")
        print(f"\n총 색상 셀: {sum(colors.values())}개")
        print("[SUCCESS] Stage 1 색상 적용됨")
        return True
    else:
        print("[FAIL] Stage 1 색상 없음")
        return False

def check_stage4():
    """Check Stage 4 report colors"""
    print("\n" + "="*80)
    print("Stage 4 색상 검증")
    print("="*80)
    
    report_dir = Path("data/processed/reports")
    files = [f for f in report_dir.glob("HVDC_*.xlsx") if "backup" not in f.name.lower()]
    
    if not files:
        print("[ERROR] 보고서 파일 없음")
        return False
    
    file_path = max(files, key=lambda p: p.stat().st_mtime)
    print(f"파일: {file_path.name}")
    
    wb = openpyxl.load_workbook(file_path)
    
    # Find target sheet
    target = None
    for name in wb.sheetnames:
        if "통합_원본데이터" in name and "Fixed" in name:
            target = name
            break
    
    if not target:
        target = wb.sheetnames[11] if len(wb.sheetnames) > 11 else wb.sheetnames[0]
    
    ws = wb[target]
    print(f"시트: {target}, {ws.max_row}행 x {ws.max_column}열")
    
    colors = Counter()
    colored_rows = set()
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(1000, ws.max_row)), start=2):
        has_color = False
        for cell in row:
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                c = str(cell.fill.fgColor.rgb)
                if c not in ['00000000', 'None']:
                    colors[c] += 1
                    has_color = True
        if has_color:
            colored_rows.add(row_idx)
    
    print(f"\n색상 결과:")
    if colors:
        for color, count in colors.most_common():
            name = "알 수 없음"
            if color == "FFFF0000":
                name = "[빨강] 시간 역전"
            elif color in ["FFFFC000", "FFC000"]:
                name = "[주황] ML 이상치-높음"
            elif color in ["FFFFFF00", "FFFF00"]:
                name = "[노랑] ML 이상치-보통"
            elif color == "FFCC99FF":
                name = "[보라] 데이터 품질"
            print(f"  {name} ({color}): {count}개")
        print(f"\n총 색상 셀: {sum(colors.values())}개")
        print(f"색상 적용 행: {len(colored_rows)}개")
        print("[SUCCESS] Stage 4 색상 적용됨")
        return True
    else:
        print("[FAIL] Stage 4 색상 없음")
        print("\n해결방법:")
        print("  python run_pipeline.py --stage 4 --stage4-visualize")
        return False

def main():
    s1_ok = check_stage1()
    s4_ok = check_stage4()
    
    print("\n" + "="*80)
    print("최종 검증 결과")
    print("="*80)
    print(f"Stage 1: {'[OK] 완료' if s1_ok else '[FAIL] 미완료'}")
    print(f"Stage 4: {'[OK] 완료' if s4_ok else '[FAIL] 미완료'}")
    
    if s1_ok and s4_ok:
        print("\n[SUCCESS] 모든 색상 작업 완료!")
    else:
        print("\n[FAIL] 일부 색상 작업 미완료")
        if not s4_ok:
            print("\nStage 4 색상 적용 명령어:")
            print("  python run_pipeline.py --stage 4 --stage4-visualize")

if __name__ == "__main__":
    main()

