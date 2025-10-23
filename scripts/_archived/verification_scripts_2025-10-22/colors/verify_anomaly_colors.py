#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HVDC 이상치 보고서 색상 적용 검증 스크립트
"""

import openpyxl
from pathlib import Path
from collections import defaultdict


def verify_anomaly_colors():
    """
    HVDC_anomaly_report.xlsx의 색상 적용을 검증합니다.
    """
    
    print("=" * 80)
    print("HVDC 이상치 보고서 색상 적용 검증")
    print("=" * 80)
    print()
    
    # 1. 파일 로드
    anomaly_report_path = Path("data/anomaly/HVDC_anomaly_report.xlsx")
    if not anomaly_report_path.exists():
        print(f"ERROR: {anomaly_report_path}를 찾을 수 없습니다.")
        return False
    
    print(f"[검증 1] 파일 로드: {anomaly_report_path}")
    wb = openpyxl.load_workbook(anomaly_report_path)
    print(f"   ✅ 파일 로드 성공")
    print(f"   전체 시트: {', '.join(wb.sheetnames)}")
    print()
    
    # 2. "통합_원본데이터_Fixed" 시트 확인
    print("[검증 2] 통합_원본데이터_Fixed 시트 확인")
    if "통합_원본데이터_Fixed" not in wb.sheetnames:
        print("   ❌ 시트를 찾을 수 없습니다.")
        return False
    
    ws = wb["통합_원본데이터_Fixed"]
    print(f"   ✅ 시트 존재 확인")
    print(f"   데이터 크기: {ws.max_row}행 × {ws.max_column}열")
    print()
    
    # 3. Case NO 컬럼 찾기
    print("[검증 3] Case NO 컬럼 찾기")
    case_col = None
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value and (str(header_value) == "Case No." or "CASE_NO" in str(header_value).upper()):
            case_col = col
            print(f"   ✅ Case NO 컬럼 발견: {col}번째 ({header_value})")
            break
    
    if not case_col:
        print("   ❌ Case NO 컬럼을 찾을 수 없습니다.")
        return False
    print()
    
    # 4. 색상 적용 확인
    print("[검증 4] 색상 적용 확인")
    
    # 색상 코드 정의 (ARGB 형식)
    color_definitions = {
        "FFFF0000": "빨강 (시간 역전)",
        "00FF0000": "빨강 (시간 역전)",
        "FFFFC000": "주황 (ML 높음)",
        "00FFC000": "주황 (ML 높음)",
        "FFFFFF00": "노랑 (ML 보통)",
        "00FFFF00": "노랑 (ML 보통)",
        "FFCC99FF": "보라 (데이터 품질)",
        "00CC99FF": "보라 (데이터 품질)",
    }
    
    colored_rows = set()
    color_counts = defaultdict(int)
    
    for row in range(2, ws.max_row + 1):
        row_colors = set()
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.start_color:
                color_code = cell.fill.start_color.rgb
                if color_code in color_definitions:
                    row_colors.add(color_code)
        
        if row_colors:
            colored_rows.add(row)
            for color in row_colors:
                color_counts[color] += 1
    
    print(f"   색상이 적용된 행: {len(colored_rows)}개")
    print()
    
    # 5. 색상별 통계
    print("[검증 5] 색상별 통계")
    if not color_counts:
        print("   ⚠️  색상이 적용된 셀을 찾을 수 없습니다.")
    else:
        for color_code, count in sorted(color_counts.items(), key=lambda x: -x[1]):
            color_name = color_definitions.get(color_code, f"알 수 없음 ({color_code})")
            print(f"   {color_name}: {count}개 셀")
    print()
    
    # 6. 샘플 케이스 확인 (처음 10개 색상 적용 행)
    print("[검증 6] 샘플 케이스 (처음 10개)")
    sample_count = 0
    for row in sorted(colored_rows)[:10]:
        case_id = ws.cell(row=row, column=case_col).value
        
        # 해당 행의 색상 확인
        row_color = None
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.start_color:
                color_code = cell.fill.start_color.rgb
                if color_code in color_definitions:
                    row_color = color_definitions[color_code]
                    break
        
        print(f"   Row {row}: Case ID = {case_id}, 색상 = {row_color}")
        sample_count += 1
    print()
    
    # 7. 색상 범례 시트 확인
    print("[검증 7] 색상 범례 시트 확인")
    if "색상_범례" not in wb.sheetnames:
        print("   ⚠️  색상_범례 시트를 찾을 수 없습니다.")
    else:
        legend_ws = wb["색상_범례"]
        print(f"   ✅ 색상_범례 시트 존재")
        print(f"   범례 데이터:")
        for row in range(1, min(legend_ws.max_row + 1, 10)):
            row_data = [legend_ws.cell(row=row, column=col).value for col in range(1, 5)]
            if any(row_data):
                print(f"      {' | '.join(str(v) if v else '' for v in row_data)}")
    print()
    
    # 8. 최종 요약
    print("=" * 80)
    print("✅ 검증 완료")
    print("=" * 80)
    print(f"시트 개수: {len(wb.sheetnames)}개")
    print(f"통합_원본데이터_Fixed: {ws.max_row}행 × {ws.max_column}열")
    print(f"색상 적용 행: {len(colored_rows)}개")
    print(f"색상 종류: {len(color_counts)}개")
    
    if len(colored_rows) > 0:
        print("\n✅ 색상 적용이 정상적으로 완료되었습니다!")
        return True
    else:
        print("\n⚠️  색상이 적용되지 않았습니다.")
        return False


if __name__ == "__main__":
    try:
        success = verify_anomaly_colors()
        exit(0 if success else 1)
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        exit(1)

