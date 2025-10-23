#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Debug color matching issue"""

import sys
import io
import json
import openpyxl
import re
from pathlib import Path
from collections import Counter

# UTF-8 encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def _norm_case(s: object) -> str:
    """Case ID 정규화: 공백/특수문자 제거 + 대문자."""
    return re.sub(r"[^A-Z0-9]", "", str(s).strip().upper()) if s is not None else ""

# Load JSON anomalies
json_path = Path("data/anomaly/HVDC_anomaly_report.json")
with open(json_path, encoding='utf-8') as f:
    anomalies = json.load(f)

print(f"JSON 이상치 총 {len(anomalies)}건")
print(f"\n타입별:")
types = Counter(a.get('Anomaly_Type') for a in anomalies)
for t, cnt in types.items():
    print(f"  {t}: {cnt}건")

# Build case index
by_case = {}
for a in anomalies:
    cid = _norm_case(a.get("Case_ID", ""))
    if cid:
        by_case.setdefault(cid, []).append(a)

print(f"\n정규화된 Case ID: {len(by_case)}개")
print(f"샘플 (처음 10개):")
for i, cid in enumerate(list(by_case.keys())[:10]):
    print(f"  {cid}")

# Load Excel
report_dir = Path("data/processed/reports")
files = [f for f in report_dir.glob("HVDC_*.xlsx") if "backup" not in f.name.lower()]
file_path = max(files, key=lambda p: p.stat().st_mtime)

print(f"\n보고서 파일: {file_path.name}")

wb = openpyxl.load_workbook(file_path)
target = None
for name in wb.sheetnames:
    if "통합_원본데이터" in name and "Fixed" in name:
        target = name
        break

if not target:
    target = wb.sheetnames[11] if len(wb.sheetnames) > 11 else wb.sheetnames[0]

ws = wb[target]
print(f"시트: {target}, {ws.max_row}행 x {ws.max_column}열")

# Find Case column
header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
case_col_idx = None
for c, name in enumerate(header, 1):
    if name and "case" in str(name).lower():
        case_col_idx = c
        print(f"\nCase 컬럼: {c}번째 ({name})")
        break

if not case_col_idx:
    print("\n[ERROR] Case NO 컬럼을 찾지 못함")
    sys.exit(1)

# Check matching
print(f"\n매칭 테스트 (처음 20행):")
matched = 0
not_matched = 0

for r in range(2, min(22, ws.max_row+1)):
    raw_id = ws.cell(row=r, column=case_col_idx).value
    cid = _norm_case(raw_id)
    
    if cid in by_case:
        matched += 1
        anoms = by_case[cid]
        print(f"  Row {r}: {raw_id} -> {cid} [OK] ({len(anoms)}개 이상치)")
    else:
        not_matched += 1
        if not_matched <= 5:  # 처음 5개만 출력
            print(f"  Row {r}: {raw_id} -> {cid} [MISS]")

print(f"\n전체 매칭 통계 (처음 1000행):")
total_matched = 0
total_checked = 0

for r in range(2, min(1001, ws.max_row+1)):
    raw_id = ws.cell(row=r, column=case_col_idx).value
    cid = _norm_case(raw_id)
    total_checked += 1
    if cid in by_case:
        total_matched += 1

print(f"  검사: {total_checked}행")
print(f"  매칭: {total_matched}행 ({total_matched/total_checked*100:.1f}%)")
print(f"  미매칭: {total_checked - total_matched}행 ({(total_checked - total_matched)/total_checked*100:.1f}%)")

if total_matched == 0:
    print("\n[CRITICAL] 매칭되는 케이스가 하나도 없습니다!")
    print("\nJSON 샘플 Case_ID (원본):")
    for i, a in enumerate(anomalies[:5]):
        print(f"  {a.get('Case_ID')}")
    
    print("\nExcel 샘플 Case No. (원본):")
    for r in range(2, 7):
        print(f"  {ws.cell(row=r, column=case_col_idx).value}")

