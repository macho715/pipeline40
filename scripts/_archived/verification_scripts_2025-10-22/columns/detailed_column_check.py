"""
CASE LIST.xlsx 파일의 HE-0214,0252 (Capacitor) 시트를 상세히 확인
특히 DHL 관련 데이터가 있는지 집중 확인
"""

import pandas as pd
import openpyxl

print("=" * 80)
print("CASE LIST.xlsx - HE-0214,0252 (Capacitor) 시트 상세 분석")
print("=" * 80)

# 1. 시트 로드
df = pd.read_excel("data/raw/CASE LIST.xlsx", sheet_name="HE-0214,0252 (Capacitor)")

print(f"\n총 행 수: {len(df)}")
print(f"총 컬럼 수: {len(df.columns)}")

# 2. 모든 컬럼 이름 출력
print("\n" + "=" * 80)
print("전체 컬럼 목록 (순서대로)")
print("=" * 80)

for i, col in enumerate(df.columns, 1):
    data_count = df[col].notna().sum()
    print(f"{i:3d}. 컬럼명: '{col}'")
    print(f"     데이터 개수: {data_count}건")

    # 컬럼명에 DHL이 포함되어 있는지 확인
    if "dhl" in str(col).lower() or "DHL" in str(col):
        print(f"     ★★★ DHL 관련 컬럼 발견! ★★★")
        if data_count > 0:
            print(f"     샘플 데이터:")
            for idx, val in enumerate(df[df[col].notna()][col].head(5), 1):
                print(f"       {idx}. {val}")

# 3. 창고 관련 컬럼 검색
print("\n" + "=" * 80)
print("창고 관련 컬럼 검색")
print("=" * 80)

warehouse_keywords = ["DHL", "DSV", "Hauler", "HAULER", "JDN", "MOSB", "AAA", "WH", "창고"]
warehouse_cols = []

for col in df.columns:
    for keyword in warehouse_keywords:
        if keyword.lower() in str(col).lower():
            warehouse_cols.append(col)
            break

print(f"발견된 창고 관련 컬럼: {len(warehouse_cols)}개")
for col in warehouse_cols:
    data_count = df[col].notna().sum()
    print(f"  - '{col}': {data_count}건")

# 4. 첫 5행 데이터 샘플 출력
print("\n" + "=" * 80)
print("첫 5행 데이터 샘플")
print("=" * 80)
print(df.head(5).to_string())

# 5. openpyxl로 직접 확인
print("\n" + "=" * 80)
print("openpyxl로 시트 직접 확인")
print("=" * 80)

wb = openpyxl.load_workbook("data/raw/CASE LIST.xlsx")
ws = wb["HE-0214,0252 (Capacitor)"]

print(f"최대 행: {ws.max_row}")
print(f"최대 열: {ws.max_column}")

print("\n첫 번째 행 (헤더) 내용:")
for col in range(1, min(ws.max_column + 1, 50)):
    cell_value = ws.cell(row=1, column=col).value
    if cell_value and "dhl" in str(cell_value).lower():
        print(f"  열 {col}: '{cell_value}' ★★★ DHL 발견!")
    elif col <= 10:  # 첫 10개만 출력
        print(f"  열 {col}: '{cell_value}'")

print("\n분석 완료")
