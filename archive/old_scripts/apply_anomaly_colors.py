#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Apply anomaly colors to the final report
"""

import json
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
from datetime import datetime


def apply_anomaly_colors():
    """Apply anomaly colors to the final report"""

    # File paths
    anomaly_json = Path("hvdc_pipeline/data/anomaly/HVDC_anomaly_report.json")
    report_file = Path(
        "hvdc_pipeline/data/processed/reports/HVDC_입고로직_종합리포트_20251019_141633_v3.0-corrected.xlsx"
    )

    # Check if files exist
    if not anomaly_json.exists():
        print(f"ERROR: Anomaly JSON file not found: {anomaly_json}")
        return False

    if not report_file.exists():
        print(f"ERROR: Report file not found: {report_file}")
        return False

    # Load anomaly data
    print("Loading anomaly data...")
    with open(anomaly_json, "r", encoding="utf-8") as f:
        anomaly_data = json.load(f)

    print(f"Found {len(anomaly_data)} anomaly records")

    # Load Excel file
    print("Loading Excel report...")
    wb = openpyxl.load_workbook(report_file)

    # Find the target sheet (first sheet)
    sheet_names = wb.sheetnames
    print(f"Available sheets: {sheet_names}")

    # Use HITACHI sheet (index 9) which contains Case No. column
    target_sheet_name = sheet_names[9]  # HITACHI_입고로직_종합리포트_Fixed
    ws = wb[target_sheet_name]
    print(f"Using sheet: {target_sheet_name}")

    # Color definitions
    colors = {
        "time_reversal": PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        ),  # Red
        "ml_high": PatternFill(
            start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"
        ),  # Orange
        "ml_medium": PatternFill(
            start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"
        ),  # Yellow
        "data_quality": PatternFill(
            start_color="FFCC99FF", end_color="FFCC99FF", fill_type="solid"
        ),  # Purple
    }

    # Case No. column is at position 11 (0-indexed: 10)
    case_col = 11
    print(f"Using Case No. column: {case_col}")

    # Apply colors based on anomaly type
    applied_count = 0
    anomaly_counts = {
        "time_reversal": 0,
        "ml_high": 0,
        "ml_medium": 0,
        "data_quality": 0,
    }

    for anomaly in anomaly_data:
        case_id = anomaly.get("Case_ID", "")
        anomaly_type = anomaly.get("Anomaly_Type", "")
        severity = anomaly.get("Severity", "")

        print(
            f"Processing anomaly: Case_ID={case_id}, Type={anomaly_type}, Severity={severity}"
        )

        # Handle special case where Case_ID is "NA" - apply to all rows with missing Case No.
        if case_id == "NA":
            print("Applying data quality color to rows with missing Case No.")
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=case_col).value
                if not cell_value or str(cell_value).strip() == "":
                    # Apply purple color to entire row for data quality issue
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = colors["data_quality"]
                    applied_count += 1
            anomaly_counts["data_quality"] += 1
        else:
            # Find matching Case ID
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=case_col).value
                if cell_value and str(cell_value).strip() == str(case_id).strip():
                    print(f"Found matching Case ID at row {row}")

                    # Apply color based on anomaly type
                    if "시간 역전" in anomaly_type or "time" in anomaly_type.lower():
                        # Time reversal: only date columns
                        for col in range(1, ws.max_column + 1):
                            header = ws.cell(row=1, column=col).value
                            if header and any(
                                keyword in str(header).lower()
                                for keyword in ["date", "날짜", "time", "시간"]
                            ):
                                ws.cell(row=row, column=col).fill = colors[
                                    "time_reversal"
                                ]
                        anomaly_counts["time_reversal"] += 1
                        applied_count += 1

                    elif "머신러닝" in anomaly_type or "ml" in anomaly_type.lower():
                        # ML anomaly: entire row based on severity
                        if severity in ["높음", "치명적", "high", "critical"]:
                            color = colors["ml_high"]
                            anomaly_counts["ml_high"] += 1
                        else:
                            color = colors["ml_medium"]
                            anomaly_counts["ml_medium"] += 1

                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = color
                        applied_count += 1

                    elif "데이터 품질" in anomaly_type or "품질" in anomaly_type:
                        # Data quality: entire row
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = colors["data_quality"]
                        anomaly_counts["data_quality"] += 1
                        applied_count += 1
                    break

    # Add color legend sheet
    print("Adding color legend...")
    legend_sheet_name = "색상 범례"
    if legend_sheet_name in wb.sheetnames:
        legend_ws = wb[legend_sheet_name]
        wb.remove(legend_ws)

    legend_ws = wb.create_sheet(legend_sheet_name)

    # Legend data
    legend_data = [
        ["색상", "의미", "적용 범위", "개수"],
        [
            "🔴 빨간색",
            "시간 역전 이상치",
            "날짜 컬럼만",
            str(anomaly_counts["time_reversal"]),
        ],
        [
            "🟠 주황색",
            "ML 이상치 (높음/치명적)",
            "전체 행",
            str(anomaly_counts["ml_high"]),
        ],
        [
            "🟡 노란색",
            "ML 이상치 (보통/낮음)",
            "전체 행",
            str(anomaly_counts["ml_medium"]),
        ],
        [
            "🟣 보라색",
            "데이터 품질 이상",
            "전체 행",
            str(anomaly_counts["data_quality"]),
        ],
        ["", "", "총 적용", str(applied_count)],
    ]

    for row_idx, row_data in enumerate(legend_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = legend_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 1:  # Header
                cell.font = openpyxl.styles.Font(bold=True)
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                bottom=openpyxl.styles.Side(style="thin"),
            )

    # Save the file with a new name to avoid permission issues
    print("Saving colored report...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = (
        report_file.parent / f"HVDC_입고로직_종합리포트_{timestamp}_colored.xlsx"
    )
    wb.save(output_file)
    print(f"Colored report saved to: {output_file}")

    print(f"SUCCESS: Applied colors to {applied_count} cases")
    print(f"Color legend added to '{legend_sheet_name}' sheet")
    print(f"Anomaly breakdown: {anomaly_counts}")

    return True


if __name__ == "__main__":
    success = apply_anomaly_colors()
    if success:
        print("Color application completed successfully!")
    else:
        print("Color application failed!")
