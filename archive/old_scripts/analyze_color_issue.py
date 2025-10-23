# -*- coding: utf-8 -*-
"""
색상 문제 상세 분석 스크립트
00000000 색상이 실제로 색상인지, 아니면 기본값인지 확인
"""

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill


def analyze_color_issue():
    """색상 문제를 상세히 분석합니다."""

    synced_file = "data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v2.9.4.xlsx"

    print("=" * 60)
    print("색상 문제 상세 분석")
    print("=" * 60)

    try:
        wb = openpyxl.load_workbook(synced_file)
        ws = wb.active

        print(f"파일: {synced_file}")
        print(f"워크시트: {ws.title}")
        print(f"크기: {ws.max_row}행 x {ws.max_column}열")

        # 색상 적용된 셀 분석
        colored_cells = []
        empty_colored_cells = []

        for row in ws.iter_rows():
            for cell in row:
                # fill 속성이 있는지 확인
                has_fill = False
                fill_type = None
                fill_color = None

                if cell.fill and cell.fill.fgColor:
                    has_fill = True
                    fill_type = cell.fill.fill_type
                    fill_color = cell.fill.fgColor.rgb

                if has_fill:
                    colored_cells.append(
                        {
                            "row": cell.row,
                            "col": cell.column,
                            "coordinate": cell.coordinate,
                            "value": cell.value,
                            "fill_type": fill_type,
                            "fill_color": fill_color,
                            "is_empty": cell.value is None
                            or str(cell.value).strip() == "",
                        }
                    )

                    if cell.value is None or str(cell.value).strip() == "":
                        empty_colored_cells.append(
                            {
                                "row": cell.row,
                                "col": cell.column,
                                "coordinate": cell.coordinate,
                                "value": cell.value,
                                "fill_type": fill_type,
                                "fill_color": fill_color,
                            }
                        )

        print(f"\n분석 결과:")
        print(f"  색상 적용된 셀: {len(colored_cells):,}")
        print(f"  빈 셀에 색상: {len(empty_colored_cells):,}")

        # fill_type 분석
        fill_types = {}
        for cell in colored_cells:
            ft = cell["fill_type"]
            if ft not in fill_types:
                fill_types[ft] = 0
            fill_types[ft] += 1

        print(f"\nFill Type 분포:")
        for ft, count in fill_types.items():
            print(f"  {ft}: {count:,}개")

        # fill_color 분석
        fill_colors = {}
        for cell in colored_cells:
            fc = cell["fill_color"]
            if fc not in fill_colors:
                fill_colors[fc] = 0
            fill_colors[fc] += 1

        print(f"\nFill Color 분포:")
        for fc, count in fill_colors.items():
            color_name = get_color_name(fc)
            print(f"  {color_name} ({fc}): {count:,}개")

        # 00000000 색상 상세 분석
        if "00000000" in fill_colors:
            print(f"\n00000000 색상 상세 분석:")
            black_cells = [c for c in colored_cells if c["fill_color"] == "00000000"]
            empty_black_cells = [c for c in black_cells if c["is_empty"]]

            print(f"  00000000 총 셀: {len(black_cells):,}")
            print(f"  빈 셀 중 00000000: {len(empty_black_cells):,}")

            # 처음 5개 셀 상세 정보
            print(f"  처음 5개 셀:")
            for i, cell in enumerate(black_cells[:5]):
                print(
                    f"    {i+1}. {cell['coordinate']} - 값: {repr(cell['value'])} - 타입: {cell['fill_type']}"
                )

        # 실제 색상이 적용된 셀 (00000000 제외)
        real_colored = [c for c in colored_cells if c["fill_color"] != "00000000"]
        real_empty_colored = [
            c for c in empty_colored_cells if c["fill_color"] != "00000000"
        ]

        print(f"\n실제 색상 적용 (00000000 제외):")
        print(f"  실제 색상 셀: {len(real_colored):,}")
        print(f"  빈 셀에 실제 색상: {len(real_empty_colored):,}")

        if len(real_empty_colored) > 0:
            print(f"\n빈 셀에 실제 색상 적용된 셀 (처음 10개):")
            for i, cell in enumerate(real_empty_colored[:10]):
                color_name = get_color_name(cell["fill_color"])
                print(
                    f"  {i+1:2d}. {cell['coordinate']} - {color_name} - 값: {repr(cell['value'])}"
                )
        else:
            print(f"\n정상: 빈 셀에 실제 색상이 적용된 셀이 없음")

        return {
            "total_colored": len(colored_cells),
            "empty_colored": len(empty_colored_cells),
            "real_colored": len(real_colored),
            "real_empty_colored": len(real_empty_colored),
            "fill_types": fill_types,
            "fill_colors": fill_colors,
        }

    except Exception as e:
        print(f"ERROR: {str(e)}")
        return None


def get_color_name(rgb):
    """RGB 색상 코드를 색상명으로 변환"""
    color_map = {
        "00000000": "기본값 (검은색)",
        "FFFFC000": "주황색 (날짜 변경)",
        "FFFFFF00": "노란색 (신규 행)",
        "FFC0C0C0": "회색",
        "FFFFFFFF": "흰색",
        "FF000000": "검은색",
    }
    return color_map.get(rgb, f"기타 ({rgb})")


if __name__ == "__main__":
    result = analyze_color_issue()

    if result:
        print(f"\n" + "=" * 60)
        print("분석 완료")
        print("=" * 60)

        if result["real_empty_colored"] > 0:
            print(f"문제: {result['real_empty_colored']}개의 빈 셀에 실제 색상 적용")
        else:
            print(f"정상: 빈 셀에 실제 색상이 적용된 셀이 없음")
            print(f"참고: 00000000은 기본값이므로 문제가 아님")
