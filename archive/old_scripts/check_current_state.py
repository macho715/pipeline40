# -*- coding: utf-8 -*-
import pandas as pd

print("=== 파일 정보 ===")
master = pd.read_excel("data/raw/Case List.xlsx")
synced = pd.read_excel(
    "data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v2.9.4.xlsx"
)

print(f"Master 행 수: {len(master)}")
print(f"Synced 행 수: {len(synced)}")
print(f"Master 컬럼 수: {len(master.columns)}")
print(f"Synced 컬럼 수: {len(synced.columns)}")

print("\n=== Master NO. 정렬 확인 ===")
master_cases = master["Case No."].dropna().tolist()[:10]
synced_cases = synced["Case No."].dropna().tolist()[:10]
print(f"Master 첫 10개: {master_cases}")
print(f"Synced 첫 10개: {synced_cases}")
print(f"일치 여부: {master_cases == synced_cases}")

# 색상 정보 확인
from openpyxl import load_workbook

print("\n=== 색상 정보 ===")
wb = load_workbook(
    "data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v2.9.4.xlsx"
)
ws = wb.active

colored_count = 0
for row in ws.iter_rows(min_row=2, max_row=min(100, ws.max_row)):
    for cell in row:
        if (
            cell.fill
            and cell.fill.start_color
            and cell.fill.start_color.rgb not in ["00000000", None]
        ):
            colored_count += 1
            if colored_count <= 5:
                print(
                    f"색상 발견: 행 {cell.row}, 컬럼 {cell.column}, 색상 {cell.fill.start_color.rgb}"
                )

print(f"첫 100행에서 색상이 적용된 셀 수: {colored_count}")
