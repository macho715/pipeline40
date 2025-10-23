#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
색상 적용 검증 스크립트
Stage 1에서 날짜 변경(주황), 신규 행(노랑) 색상이 올바르게 적용되었는지 확인
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
from pathlib import Path


def verify_colors_applied():
    """색상 적용 상태를 검증합니다."""

    print("=== 색상 적용 검증 ===")

    try:
        # 파일 경로 설정
        synced_file = (
            "data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v2.9.4.xlsx"
        )

        # 파일 존재 확인
        if not Path(synced_file).exists():
            print(f"ERROR: Synced 파일이 존재하지 않습니다: {synced_file}")
            return False

        print(f"파일 로딩 중: {synced_file}")

        # Excel 파일 로드
        wb = load_workbook(synced_file)
        ws = wb.active

        print(f"워크시트 정보: {ws.max_row}행, {ws.max_column}컬럼")

        # 색상 카운터 초기화
        orange_count = 0  # FFC000 (날짜 변경)
        yellow_count = 0  # FFFF00 (신규 행)
        other_colors = {}  # 기타 색상들

        # 헤더 행 제외하고 모든 셀 검사
        print("색상 검사 중...")

        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # 셀의 배경색 확인
                if hasattr(cell.fill, "start_color") and cell.fill.start_color.rgb:
                    color = cell.fill.start_color.rgb

                    if color == "00FFC000":  # 주황색 (날짜 변경)
                        orange_count += 1
                    elif color == "00FFFF00":  # 노랑색 (신규 행)
                        yellow_count += 1
                    elif color != "00000000":  # 투명이 아닌 다른 색상
                        if color in other_colors:
                            other_colors[color] += 1
                        else:
                            other_colors[color] = 1

        # 결과 출력
        print(f"\n색상 적용 결과:")
        print(f"주황색 (날짜 변경): {orange_count}개 셀")
        print(f"노랑색 (신규 행): {yellow_count}개 셀")

        if other_colors:
            print(f"기타 색상:")
            for color, count in other_colors.items():
                print(f"   {color}: {count}개 셀")

        # 예상 값과 비교
        print(f"\n검증 결과:")

        # 주황색 검증 (날짜 업데이트)
        if orange_count >= 1000:  # 최소 1000개 이상
            print(f"PASS 주황색 (날짜 변경): {orange_count}개 - 정상 범위")
        else:
            print(f"WARNING 주황색 (날짜 변경): {orange_count}개 - 예상보다 적음")

        # 노랑색 검증 (신규 행)
        if yellow_count >= 100:  # 최소 100개 이상
            print(f"PASS 노랑색 (신규 행): {yellow_count}개 - 정상 범위")
        else:
            print(f"WARNING 노랑색 (신규 행): {yellow_count}개 - 예상보다 적음")

        # 전체 평가
        total_colored = orange_count + yellow_count
        if total_colored > 0:
            print(f"PASS 총 {total_colored}개 셀에 색상이 적용되었습니다")
            return True
        else:
            print(f"ERROR 색상이 적용된 셀이 없습니다")
            return False

    except Exception as e:
        print(f"ERROR 오류 발생: {str(e)}")
        return False


if __name__ == "__main__":
    success = verify_colors_applied()
    sys.exit(0 if success else 1)
