# -*- coding: utf-8 -*-
"""
Stage 1 색상 표시 검증 스크립트
빈 공간(데이터 없는 셀)에 색상이 잘못 적용된 셀을 찾아 분석합니다.
"""

import openpyxl
import pandas as pd
from openpyxl.styles import Color
import sys
import os


def diagnose_color_in_empty_cells():
    """빈 셀에 색상이 적용된 셀들을 찾아 분석합니다."""

    # 파일 경로 설정
    synced_file = "data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v2.9.4.xlsx"

    if not os.path.exists(synced_file):
        print(f"ERROR: 파일을 찾을 수 없습니다: {synced_file}")
        return

    print("=" * 60)
    print("Stage 1 색상 표시 검증 - 빈 셀 색상 분석")
    print("=" * 60)

    try:
        # Excel 파일 로드
        print(f"파일 로딩 중: {synced_file}")
        wb = openpyxl.load_workbook(synced_file)
        ws = wb.active

        print(f"워크시트 활성화: {ws.title}")
        print(f"데이터 범위: {ws.max_row}행 x {ws.max_column}열")

        # 색상이 적용된 빈 셀 찾기
        empty_colored_cells = []
        colored_cells = []
        total_cells = 0

        print("\n셀 분석 중...")

        for row in ws.iter_rows():
            for cell in row:
                total_cells += 1

                # 색상이 적용된 셀 찾기
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    color_rgb = cell.fill.fgColor.rgb
                    colored_cells.append(
                        {
                            "row": cell.row,
                            "col": cell.column,
                            "coordinate": cell.coordinate,
                            "color": color_rgb,
                            "value": cell.value,
                            "value_type": type(cell.value).__name__,
                        }
                    )

                    # 빈 셀인지 확인
                    is_empty = False
                    if cell.value is None:
                        is_empty = True
                    elif isinstance(cell.value, str) and cell.value.strip() == "":
                        is_empty = True
                    elif pd.isna(cell.value):
                        is_empty = True

                    if is_empty:
                        empty_colored_cells.append(
                            {
                                "row": cell.row,
                                "col": cell.column,
                                "coordinate": cell.coordinate,
                                "color": color_rgb,
                                "value": cell.value,
                                "value_type": type(cell.value).__name__,
                            }
                        )

        # 결과 분석
        print(f"\n분석 완료:")
        print(f"  전체 셀 수: {total_cells:,}")
        print(f"  색상 적용된 셀: {len(colored_cells):,}")
        print(f"  빈 셀에 색상 적용: {len(empty_colored_cells):,}")

        if len(empty_colored_cells) > 0:
            print(f"\n문제 발견: {len(empty_colored_cells)}개의 빈 셀에 색상이 적용됨")

            # 색상별 분류
            color_stats = {}
            for cell in empty_colored_cells:
                color = cell["color"]
                if color not in color_stats:
                    color_stats[color] = 0
                color_stats[color] += 1

            print(f"\n색상별 빈 셀 분포:")
            for color, count in color_stats.items():
                color_name = get_color_name(color)
                print(f"  {color_name} ({color}): {count}개")

            # 처음 10개 빈 셀 상세 정보
            print(f"\n빈 셀 색상 적용 상세 (처음 10개):")
            for i, cell in enumerate(empty_colored_cells[:10]):
                print(
                    f"  {i+1:2d}. {cell['coordinate']} - {get_color_name(cell['color'])} - 값: {repr(cell['value'])}"
                )

            if len(empty_colored_cells) > 10:
                print(f"  ... 외 {len(empty_colored_cells) - 10}개")

        else:
            print(f"\n정상: 빈 셀에 색상이 적용된 셀이 없음")

        # 색상 적용 패턴 분석
        print(f"\n색상 적용 패턴 분석:")
        analyze_color_patterns(colored_cells)

        # 빈 셀 패턴 분석
        if len(empty_colored_cells) > 0:
            print(f"\n빈 셀 색상 패턴 분석:")
            analyze_empty_cell_patterns(empty_colored_cells)

        return {
            "total_cells": total_cells,
            "colored_cells": len(colored_cells),
            "empty_colored_cells": len(empty_colored_cells),
            "empty_cells_list": empty_colored_cells,
            "color_stats": color_stats if len(empty_colored_cells) > 0 else {},
        }

    except Exception as e:
        print(f"ERROR: 분석 중 오류 발생: {str(e)}")
        return None


def get_color_name(rgb):
    """RGB 색상 코드를 색상명으로 변환"""
    color_map = {
        "FFFFC000": "주황색 (날짜 변경)",
        "FFFFFF00": "노란색 (신규 행)",
        "FFC0C0C0": "회색",
        "FFFFFFFF": "흰색",
        "FF000000": "검은색",
    }
    return color_map.get(rgb, f"기타 ({rgb})")


def analyze_color_patterns(colored_cells):
    """색상 적용 패턴을 분석합니다."""
    if not colored_cells:
        return

    # 색상별 통계
    color_stats = {}
    for cell in colored_cells:
        color = cell["color"]
        if color not in color_stats:
            color_stats[color] = 0
        color_stats[color] += 1

    print(f"  전체 색상 적용 분포:")
    for color, count in color_stats.items():
        color_name = get_color_name(color)
        percentage = (count / len(colored_cells)) * 100
        print(f"    {color_name}: {count:,}개 ({percentage:.1f}%)")

    # 행별 색상 적용 패턴
    row_colors = {}
    for cell in colored_cells:
        row = cell["row"]
        if row not in row_colors:
            row_colors[row] = []
        row_colors[row].append(cell["color"])

    # 색상이 많이 적용된 행들
    multi_color_rows = {
        row: colors for row, colors in row_colors.items() if len(set(colors)) > 1
    }
    if multi_color_rows:
        print(f"  다중 색상 행: {len(multi_color_rows)}개")
        for row in sorted(multi_color_rows.keys())[:5]:
            colors = set(multi_color_rows[row])
            print(f"    행 {row}: {', '.join([get_color_name(c) for c in colors])}")


def analyze_empty_cell_patterns(empty_colored_cells):
    """빈 셀에 색상이 적용된 패턴을 분석합니다."""
    if not empty_colored_cells:
        return

    # 행별 분포
    row_distribution = {}
    for cell in empty_colored_cells:
        row = cell["row"]
        if row not in row_distribution:
            row_distribution[row] = 0
        row_distribution[row] += 1

    print(f"  빈 셀 색상이 많은 행 (상위 10개):")
    sorted_rows = sorted(row_distribution.items(), key=lambda x: x[1], reverse=True)
    for row, count in sorted_rows[:10]:
        print(f"    행 {row}: {count}개")

    # 연속된 행 패턴 확인
    rows_with_empty_colored = sorted(row_distribution.keys())
    consecutive_groups = []
    current_group = [rows_with_empty_colored[0]]

    for i in range(1, len(rows_with_empty_colored)):
        if rows_with_empty_colored[i] == rows_with_empty_colored[i - 1] + 1:
            current_group.append(rows_with_empty_colored[i])
        else:
            if len(current_group) > 1:
                consecutive_groups.append(current_group)
            current_group = [rows_with_empty_colored[i]]

    if len(current_group) > 1:
        consecutive_groups.append(current_group)

    if consecutive_groups:
        print(f"  연속된 행 그룹: {len(consecutive_groups)}개")
        for group in consecutive_groups[:5]:
            print(f"    행 {group[0]}-{group[-1]}: {len(group)}개 행")


if __name__ == "__main__":
    result = diagnose_color_in_empty_cells()

    if result:
        print(f"\n" + "=" * 60)
        print("진단 완료")
        print("=" * 60)

        if result["empty_colored_cells"] > 0:
            print(f"문제 확인: {result['empty_colored_cells']}개의 빈 셀에 색상 적용")
            print(
                "다음 단계: data_synchronizer_v29.py의 _apply_colors() 로직 검토 필요"
            )
        else:
            print(f"정상: 빈 셀에 색상이 적용된 셀이 없음")
    else:
        print("진단 실패")
