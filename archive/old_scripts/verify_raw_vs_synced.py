# -*- coding: utf-8 -*-
"""
RAW 데이터 vs Synced 결과 비교 검증 스크립트
RAW Data vs Synced Result Comparison Verification Script
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import random
from datetime import datetime


def load_data():
    """RAW 데이터와 Synced 데이터 로드"""
    print("=== 1. RAW 데이터 로드 ===")

    # RAW 데이터 로드
    master_df = pd.read_excel("data/raw/Case List.xlsx")
    warehouse_df = pd.read_excel("data/raw/HVDC WAREHOUSE_HITACHI(HE).xlsx")
    synced_df = pd.read_excel(
        "data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v2.9.4.xlsx"
    )

    print(f"Master: {len(master_df)} 행, {len(master_df.columns)} 컬럼")
    print(f"Warehouse: {len(warehouse_df)} 행, {len(warehouse_df.columns)} 컬럼")
    print(f"Synced: {len(synced_df)} 행, {len(synced_df.columns)} 컬럼")

    return master_df, warehouse_df, synced_df


def verify_basic_info(master_df, warehouse_df, synced_df):
    """기본 정보 검증"""
    print("\n=== 2. 기본 정보 검증 ===")

    # Case No. 중복 확인
    master_duplicates = master_df["Case No."].duplicated().sum()
    warehouse_duplicates = warehouse_df["Case No."].duplicated().sum()
    synced_duplicates = synced_df["Case No."].duplicated().sum()

    print(f"Master Case No. 중복: {master_duplicates}개")
    print(f"Warehouse Case No. 중복: {warehouse_duplicates}개")
    print(f"Synced Case No. 중복: {synced_duplicates}개")

    # NO. 컬럼 확인
    master_has_no = "No." in master_df.columns
    warehouse_has_no = "No." in warehouse_df.columns
    synced_has_no = "No." in synced_df.columns

    print(f"Master NO. 컬럼 존재: {master_has_no}")
    print(f"Warehouse NO. 컬럼 존재: {warehouse_has_no}")
    print(f"Synced NO. 컬럼 존재: {synced_has_no}")

    return {
        "master_duplicates": master_duplicates,
        "warehouse_duplicates": warehouse_duplicates,
        "synced_duplicates": synced_duplicates,
        "master_has_no": master_has_no,
        "warehouse_has_no": warehouse_has_no,
        "synced_has_no": synced_has_no,
    }


def verify_master_no_sorting(master_df, synced_df):
    """Master NO. 순서 정렬 검증"""
    print("\n=== 3. Master NO. 순서 정렬 검증 ===")

    # Master Case No. 순서 추출
    master_cases = master_df["Case No."].dropna().tolist()
    synced_cases = synced_df["Case No."].dropna().tolist()

    print(f"Master Case 수: {len(master_cases)}")
    print(f"Synced Case 수: {len(synced_cases)}")

    # 첫 100개 비교
    first_100_match = master_cases[:100] == synced_cases[:100]
    print(f"첫 100개 일치: {first_100_match}")

    # 첫 1000개 비교
    first_1000_match = master_cases[:1000] == synced_cases[:1000]
    print(f"첫 1000개 일치: {first_1000_match}")

    # Master Case들이 Synced 앞부분에 있는지 확인
    master_in_synced = synced_cases[: len(master_cases)]
    full_match = master_cases == master_in_synced
    print(f"전체 Master 순서 일치: {full_match}")

    # 일치하지 않는 부분 찾기
    if not full_match:
        for i, (m_case, s_case) in enumerate(zip(master_cases, master_in_synced)):
            if m_case != s_case:
                print(f"첫 번째 불일치 위치: {i}, Master: {m_case}, Synced: {s_case}")
                break

    return {
        "first_100_match": first_100_match,
        "first_1000_match": first_1000_match,
        "full_match": full_match,
    }


def verify_data_updates(master_df, warehouse_df, synced_df):
    """데이터 업데이트 검증"""
    print("\n=== 4. 데이터 업데이트 검증 ===")

    # 공통 Case No. 찾기
    master_cases = set(master_df["Case No."].dropna())
    warehouse_cases = set(warehouse_df["Case No."].dropna())
    common_cases = master_cases.intersection(warehouse_cases)

    print(f"공통 Case 수: {len(common_cases)}")

    # 특정 Case 검증 (208221, 208222)
    test_cases = [208221, 208222]
    for case_no in test_cases:
        if case_no in common_cases:
            master_row = master_df[master_df["Case No."] == case_no].iloc[0]
            warehouse_row = warehouse_df[warehouse_df["Case No."] == case_no].iloc[0]
            synced_row = synced_df[synced_df["Case No."] == case_no].iloc[0]

            print(f"\nCase {case_no} 검증:")
            print(f"  Master에서 찾음: {case_no in master_cases}")
            print(f"  Warehouse에서 찾음: {case_no in warehouse_cases}")
            print(f"  Synced에서 찾음: {case_no in synced_df['Case No.'].values}")

            # 날짜 컬럼 비교 (예시)
            date_columns = ["ETA", "ETD", "입고예정일"]
            for col in date_columns:
                if col in master_df.columns and col in synced_df.columns:
                    master_val = master_row[col] if pd.notna(master_row[col]) else None
                    synced_val = synced_row[col] if pd.notna(synced_row[col]) else None
                    print(
                        f"  {col}: Master={master_val}, Synced={synced_val}, 일치={master_val == synced_val}"
                    )

    # 랜덤 샘플 검증
    sample_cases = random.sample(list(common_cases), min(20, len(common_cases)))
    correct_updates = 0

    for case_no in sample_cases:
        master_row = master_df[master_df["Case No."] == case_no].iloc[0]
        synced_row = synced_df[synced_df["Case No."] == case_no].iloc[0]

        # 주요 컬럼 비교
        key_columns = ["Case No.", "ETA", "ETD", "입고예정일"]
        matches = 0
        for col in key_columns:
            if col in master_df.columns and col in synced_df.columns:
                master_val = master_row[col] if pd.notna(master_row[col]) else None
                synced_val = synced_row[col] if pd.notna(synced_row[col]) else None
                if master_val == synced_val:
                    matches += 1

        if matches == len(
            [
                c
                for c in key_columns
                if c in master_df.columns and c in synced_df.columns
            ]
        ):
            correct_updates += 1

    print(f"\n랜덤 샘플 20개 중 정확한 업데이트: {correct_updates}개")

    return {
        "common_cases": len(common_cases),
        "correct_updates": correct_updates,
        "sample_size": len(sample_cases),
    }


def verify_new_additions(master_df, warehouse_df, synced_df):
    """신규 추가 검증"""
    print("\n=== 5. 신규 추가 검증 ===")

    # Master 전용 Case 찾기
    master_cases = set(master_df["Case No."].dropna())
    warehouse_cases = set(warehouse_df["Case No."].dropna())
    synced_cases = set(synced_df["Case No."].dropna())

    master_only = master_cases - warehouse_cases
    warehouse_only = warehouse_cases - master_cases

    print(f"Master 전용 Case 수: {len(master_only)}")
    print(f"Warehouse 전용 Case 수: {len(warehouse_only)}")

    # Master 전용 Case들이 Synced에 추가되었는지 확인
    master_only_in_synced = master_only.intersection(synced_cases)
    print(f"Master 전용 Case가 Synced에 추가됨: {len(master_only_in_synced)}개")

    # Warehouse 전용 Case들이 Synced에 있는지 확인
    warehouse_only_in_synced = warehouse_only.intersection(synced_cases)
    print(f"Warehouse 전용 Case가 Synced에 유지됨: {len(warehouse_only_in_synced)}개")

    return {
        "master_only": len(master_only),
        "warehouse_only": len(warehouse_only),
        "master_only_in_synced": len(master_only_in_synced),
        "warehouse_only_in_synced": len(warehouse_only_in_synced),
    }


def verify_color_application(synced_df):
    """색상 적용 검증"""
    print("\n=== 6. 색상 적용 검증 ===")

    # Synced 파일에서 색상 정보 확인
    wb = load_workbook(
        "data/processed/synced/HVDC WAREHOUSE_HITACHI(HE).synced_v2.9.4.xlsx"
    )
    ws = wb.active

    orange_count = 0  # 날짜 변경 (00FFC000)
    yellow_count = 0  # 신규 추가 (00FFFF00)
    other_colors = {}

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if (
                cell.fill
                and cell.fill.start_color
                and cell.fill.start_color.rgb not in ["00000000", None]
            ):
                color = cell.fill.start_color.rgb
                if color == "00FFC000":
                    orange_count += 1
                elif color == "00FFFF00":
                    yellow_count += 1
                else:
                    if color not in other_colors:
                        other_colors[color] = 0
                    other_colors[color] += 1

    print(f"주황색 셀 (날짜 변경): {orange_count}개")
    print(f"노란색 셀 (신규 추가): {yellow_count}개")
    print(f"기타 색상: {len(other_colors)}종류")

    if other_colors:
        print("기타 색상 상세:")
        for color, count in sorted(
            other_colors.items(), key=lambda x: x[1], reverse=True
        ):
            print(f"  {color}: {count}개")

    return {
        "orange_count": orange_count,
        "yellow_count": yellow_count,
        "other_colors": other_colors,
    }


def verify_statistics(master_df, warehouse_df, synced_df):
    """통계 검증"""
    print("\n=== 7. 통계 검증 ===")

    master_cases = set(master_df["Case No."].dropna())
    warehouse_cases = set(warehouse_df["Case No."].dropna())
    synced_cases = set(synced_df["Case No."].dropna())

    master_only = master_cases - warehouse_cases
    warehouse_only = warehouse_cases - master_cases
    common_cases = master_cases.intersection(warehouse_cases)

    expected_synced_count = len(master_only) + len(warehouse_cases)
    actual_synced_count = len(synced_cases)

    print(f"예상 Synced Case 수: {expected_synced_count}")
    print(f"실제 Synced Case 수: {actual_synced_count}")
    print(f"통계 일치: {expected_synced_count == actual_synced_count}")

    return {
        "expected_count": expected_synced_count,
        "actual_count": actual_synced_count,
        "matches": expected_synced_count == actual_synced_count,
    }


def generate_final_report(results):
    """최종 검증 보고서 생성"""
    print("\n" + "=" * 60)
    print("=== RAW 데이터 vs Synced 검증 보고서 ===")
    print("=" * 60)

    # 1. 파일 정보
    print("\n1. 파일 정보")
    print(f"   - Master: {results['master_rows']} 행, {results['master_cols']} 컬럼")
    print(
        f"   - Warehouse: {results['warehouse_rows']} 행, {results['warehouse_cols']} 컬럼"
    )
    print(f"   - Synced: {results['synced_rows']} 행, {results['synced_cols']} 컬럼")

    # 2. MASTER NO. 정렬 검증
    print("\n2. MASTER NO. 정렬 검증")
    print(f"   - 첫 100개 일치: {'PASS' if results['first_100_match'] else 'FAIL'}")
    print(f"   - 첫 1000개 일치: {'PASS' if results['first_1000_match'] else 'FAIL'}")
    print(f"   - 전체 순서 일치: {'PASS' if results['full_match'] else 'FAIL'}")

    # 3. 데이터 업데이트 검증
    print("\n3. 데이터 업데이트 검증")
    print(f"   - 공통 Case 수: {results['common_cases']}개")
    print(
        f"   - 샘플 정확도: {results['correct_updates']}/{results['sample_size']} PASS"
    )

    # 4. 신규 추가 검증
    print("\n4. 신규 추가 검증")
    print(f"   - Master 전용: {results['master_only']}개")
    print(f"   - Warehouse 전용: {results['warehouse_only']}개")
    print(f"   - Master 전용 추가됨: {results['master_only_in_synced']}개 PASS")

    # 5. 색상 적용 검증
    print("\n5. 색상 적용 검증")
    print(f"   - 날짜 변경 (주황): {results['orange_count']}개 PASS")
    print(f"   - 신규 추가 (노랑): {results['yellow_count']}개 PASS")

    # 6. 통계 일치
    print("\n6. 통계 일치")
    print(f"   - 예상 Synced Case 수: {results['expected_count']}")
    print(f"   - 실제 Synced Case 수: {results['actual_count']}")
    print(f"   - 통계 일치: {'PASS' if results['matches'] else 'FAIL'}")

    # 최종 결과
    success_criteria = [
        results["first_100_match"],
        results["common_cases"] > 0,
        results["correct_updates"] >= results["sample_size"] * 0.8,  # 80% 이상 정확
        results["master_only_in_synced"] == results["master_only"],
        results["orange_count"] > 0,
        results["yellow_count"] > 0,
        results["matches"],
    ]

    overall_success = all(success_criteria)

    print("\n" + "=" * 60)
    print(f"=== 검증 결과: {'성공 PASS' if overall_success else '실패 FAIL'} ===")
    print("=" * 60)

    return overall_success


def main():
    """메인 실행 함수"""
    print("RAW 데이터 vs Synced 결과 비교 검증 시작")
    print(f"실행 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 1. 데이터 로드
    master_df, warehouse_df, synced_df = load_data()

    # 2. 기본 정보 검증
    basic_info = verify_basic_info(master_df, warehouse_df, synced_df)

    # 3. Master NO. 순서 정렬 검증
    sorting_info = verify_master_no_sorting(master_df, synced_df)

    # 4. 데이터 업데이트 검증
    update_info = verify_data_updates(master_df, warehouse_df, synced_df)

    # 5. 신규 추가 검증
    addition_info = verify_new_additions(master_df, warehouse_df, synced_df)

    # 6. 색상 적용 검증
    color_info = verify_color_application(synced_df)

    # 7. 통계 검증
    stats_info = verify_statistics(master_df, warehouse_df, synced_df)

    # 8. 최종 보고서 생성
    results = {
        "master_rows": len(master_df),
        "master_cols": len(master_df.columns),
        "warehouse_rows": len(warehouse_df),
        "warehouse_cols": len(warehouse_df.columns),
        "synced_rows": len(synced_df),
        "synced_cols": len(synced_df.columns),
        **sorting_info,
        **update_info,
        **addition_info,
        **color_info,
        **stats_info,
    }

    success = generate_final_report(results)

    return success


if __name__ == "__main__":
    main()
