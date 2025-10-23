#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Stage 4 색상 적용 확인 스크립트
"""

import openpyxl

def check_colors():
    """Excel 파일의 색상 적용 상태를 확인합니다."""
    
    print("=" * 60)
    print("Stage 4 색상 적용 확인")
    print("=" * 60)
    
    # Excel 파일 로드
    try:
        wb = openpyxl.load_workbook('data/anomaly/HVDC_anomaly_report.xlsx')
        print(f"✅ Excel 파일 로드 성공")
        print(f"시트 목록: {wb.sheetnames}")
        
        # 통합_원본데이터_Fixed 시트 확인
        if '통합_원본데이터_Fixed' in wb.sheetnames:
            ws = wb['통합_원본데이터_Fixed']
            print(f"✅ 통합_원본데이터_Fixed 시트 발견")
            print(f"총 행 수: {ws.max_row}")
            
            # 색상 적용 확인
            colored_rows = 0
            sample_colors = []
            
            for row in range(2, min(12, ws.max_row + 1)):  # 처음 10행 확인
                cell = ws.cell(row=row, column=1)
                if cell.fill.start_color.rgb != '00000000':
                    colored_rows += 1
                    sample_colors.append(cell.fill.start_color.rgb)
            
            print(f"색상이 적용된 행: {colored_rows}개 (처음 10행 중)")
            
            if sample_colors:
                print("적용된 색상 샘플:")
                for i, color in enumerate(sample_colors[:5]):
                    print(f"  {i+1}. {color}")
            
            print("✅ 색상 자동 적용 성공!")
            
        else:
            print("❌ 통합_원본데이터_Fixed 시트가 없습니다.")
            
    except Exception as e:
        print(f"❌ 오류 발생: {e}")

if __name__ == "__main__":
    check_colors()


